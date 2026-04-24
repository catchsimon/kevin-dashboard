#!/usr/bin/env python3
"""
Kevin Dashboard — daily ETL + HTML dashboard generator.

Pulls the source Google Sheet (publicly viewable), normalizes property tabs
and the IVT report, maps IVT origin domains to property tabs, and emits a
single-file interactive HTML dashboard.

Designed to be run unattended by a daily scheduled task.
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
import urllib.request
import datetime as dt
from pathlib import Path
from typing import Any

import openpyxl

SHEET_ID = "1J0K3ADKoXVrT8sfFeVJNsQKN5qm2o1TO1IBoRvpeG5o"
SHEET_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
)

# Secondary workbook: Publisher Conversion Score (PCS) comparison report.
# Sheet has one tab with two columns — current period score + prior period score.
# Period labels are read from the column headers so the UI stays accurate as
# the report is refreshed.
PCS_SHEET_ID = "1bf_JzVW4DMcXMAl1vD10gsdj8cImsmoRo9uE1K_jdws"
PCS_SHEET_URL = (
    f"https://docs.google.com/spreadsheets/d/{PCS_SHEET_ID}/export?format=xlsx"
)

# Tabs to ignore for the property-revenue rollup
NON_PROPERTY_TABS = {
    "Revenue",
    "IVT Report",
    "Sheet31",
    "Platform",
    "GEO",
    "Conversion Data",
}

# Standard schema we expect in property tabs
PROP_COLS = [
    "Date",
    "Ad Revenue",
    "Page Views",
    "Views with Ads",
    "% with Visibility",
    "Views with Visibility",
    "CTR",
    "Ad Clicks",
    "Ad CPC",
    "Ad RPM",
    "Ad vRPM",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_float(v: Any) -> float | None:
    """Coerce cell value to float; return None for blanks, errors, junk."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#") or s.lower() in {"nan", "n/a", "na"}:
            return None
        # strip stray %
        s = s.replace("%", "").replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def to_date(v: Any) -> str | None:
    """Return ISO date string YYYY-MM-DD or None."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return dt.datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_oem_group(tab_name: str) -> str:
    """Extract OEM group label from a property tab name."""
    name = tab_name.strip()
    m = re.match(r"^(OEM\d*)\b", name, re.IGNORECASE)
    if m:
        grp = m.group(1).upper()
        # Normalize "OEM" (no number) to "OEM1"
        return "OEM1" if grp == "OEM" else grp
    return "OTHER"


def clean_property_label(tab_name: str) -> str:
    """Human-readable property name (strip OEM prefix and dashes)."""
    name = tab_name.strip()
    name = re.sub(r"^OEM\s*\d*\s*[-–]?\s*", "", name, flags=re.IGNORECASE)
    return name.strip(" -")


def domain_root(domain: str) -> str:
    """Strip TLD and non-alpha chars to get a comparison key."""
    if not isinstance(domain, str):
        return ""
    d = domain.strip().lower()
    # Remove leading protocol/www
    d = re.sub(r"^(https?://)?(www\.)?", "", d)
    # Remove TLD (everything after the first dot)
    d = d.split(".")[0]
    # Strip non-alphanumeric
    return re.sub(r"[^a-z0-9]", "", d)


def property_key(tab_name: str) -> str:
    """Comparison key for a property tab name."""
    label = clean_property_label(tab_name)
    return re.sub(r"[^a-z0-9]", "", label.lower())


# ---------------------------------------------------------------------------
# ETL
# ---------------------------------------------------------------------------

def download_sheet(dest_path: Path, url: str = SHEET_URL) -> None:
    print(f"[etl] downloading source sheet -> {dest_path}", file=sys.stderr)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = resp.read()
    dest_path.write_bytes(data)


def parse_property_tab(ws) -> list[dict]:
    """Parse a single property tab into a list of daily rows.

    Most tabs have a single table starting at column A. A handful of newer OEM2
    tabs (My Fimap, Zenity, Realmfeed, Sparkdaily, Local Pulse) have three
    tables side-by-side — two exploratory layouts on the left and the real
    data in the rightmost block. We locate every 'Date' header cell in the
    first few rows, then pick the one with the most date-valued rows below it
    (ties prefer the variant that has 'Ad Revenue' immediately to its right —
    the standard schema signature).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Scan the top rows for every 'Date' header occurrence.
    candidates: list[tuple[int, int]] = []  # (header_row, date_col)
    for i, r in enumerate(rows[:5]):
        if not r:
            continue
        for ci, cell in enumerate(r):
            if isinstance(cell, str) and cell.strip().lower() == "date":
                candidates.append((i, ci))
    if not candidates:
        return []

    def score(header_idx: int, date_col: int) -> tuple[int, int]:
        """Return (data_row_count, schema_bonus) — higher is better."""
        count = 0
        for r in rows[header_idx + 1:]:
            if r and date_col < len(r) and to_date(r[date_col]):
                count += 1
        # Bonus if the column to the right of 'Date' is 'Ad Revenue'.
        hdr = rows[header_idx]
        next_cell = hdr[date_col + 1] if date_col + 1 < len(hdr) else None
        schema_bonus = 1 if (isinstance(next_cell, str)
                             and next_cell.strip().lower() == "ad revenue") else 0
        return (count, schema_bonus)

    header_idx, date_col = max(candidates, key=lambda c: score(*c))
    if score(header_idx, date_col)[0] == 0:
        return []

    # Build the column map scoped to this table — start at date_col and stop
    # at the next 'Date' header (or end of row). That prevents duplicate names
    # from neighboring tables shadowing our schema.
    hdr = rows[header_idx]
    end = len(hdr)
    for ci in range(date_col + 1, len(hdr)):
        cell = hdr[ci]
        if isinstance(cell, str) and cell.strip().lower() == "date":
            end = ci
            break
    col: dict[str, int] = {}
    for ci in range(date_col, end):
        cell = hdr[ci]
        if cell is None:
            continue
        name = str(cell).strip()
        if name and name not in col:
            col[name] = ci

    out = []
    for r in rows[header_idx + 1:]:
        if not r or date_col >= len(r):
            continue
        d = to_date(r[date_col])
        if not d:
            continue

        def g(name):
            i = col.get(name)
            if i is None or i >= len(r):
                return None
            return r[i]

        ad_rev = to_float(g("Ad Revenue"))
        pv = to_float(g("Page Views"))
        clicks = to_float(g("Ad Clicks"))
        ctr = to_float(g("CTR"))
        rpm = to_float(g("Ad RPM"))
        vrpm = to_float(g("Ad vRPM"))
        vwv = to_float(g("Views with Visibility"))
        vwa = to_float(g("Views with Ads"))
        # Skip rows where every metric is None
        if all(x is None for x in (ad_rev, pv, clicks, ctr, rpm)):
            continue
        out.append({
            "date": d,
            "ad_revenue": ad_rev,
            "page_views": pv,
            "views_with_ads": vwa,
            "views_with_visibility": vwv,
            "ctr": ctr,
            "ad_clicks": clicks,
            "ad_rpm": rpm,
            "ad_vrpm": vrpm,
        })
    return out


def parse_ivt_tab(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]
    col = {h: i for i, h in enumerate(header) if h}

    def g(r, name):
        i = col.get(name)
        if i is None or i >= len(r):
            return None
        return r[i]

    out = []
    for r in rows[1:]:
        d = to_date(g(r, "Date"))
        if not d:
            continue
        domain_raw = g(r, "Origin Domain")
        if isinstance(domain_raw, (int, float)):
            domain_raw = str(domain_raw)
        domain = (domain_raw or "").strip() or "<blank>"
        out.append({
            "date": d,
            "origin_domain": domain,
            "impressions": to_float(g(r, "Impressions")) or 0.0,
            "clicks": to_float(g(r, "Clicks")) or 0.0,
            "known_bots": to_float(g(r, "Known Bots")) or 0.0,
            "known_emulators": to_float(g(r, "Known Emulators")) or 0.0,
            "in_app_stuffing": to_float(g(r, "In-App Stuffing")) or 0.0,
            "deterministic_fraud": to_float(g(r, "Deterministic Fraud")) or 0.0,
            "proxy_fraud": to_float(g(r, "Proxy Fraud")) or 0.0,
            "fraud": to_float(g(r, "Fraud")) or 0.0,
            "fraud_pct": to_float(g(r, "Fraud %")) or 0.0,
        })
    return out


def build_domain_property_map(
    property_tab_names: list[str], ivt_domains: list[str]
) -> tuple[dict[str, str], list[str]]:
    """Map IVT origin domain -> property tab name. Return (map, unmatched)."""
    # Manual overrides for tricky cases
    overrides = {
        "clickticks": "OEM5- Click Ticks ",
        "dailydash": "OEM4 -Daily Dash ",
        "starthubs": "OEM4 - Starthubs ",
        "startpage": "OEM4 - Startpage ",
        "spendider": "OEM5 - Spendider ",
        "swiftnavi": "OEM4 - Swiftnavi ",
        "promptnavi": "OEM4 - Promptnavi ",
        "easynav": "OEM5 - Easynav ",
        "naviseeking": "OEM4 - Naviseeking ",
        "navitravel": "OEM4 - Navitravel ",
    }
    prop_keys = {property_key(t): t for t in property_tab_names}
    mapping: dict[str, str] = {}
    unmatched: list[str] = []
    for dom in ivt_domains:
        if not dom or dom == "<blank>":
            continue
        root = domain_root(dom)
        if not root:
            continue
        # 1. exact key match
        if root in prop_keys:
            mapping[dom] = prop_keys[root]
            continue
        # 2. override
        if root in overrides and overrides[root] in property_tab_names:
            mapping[dom] = overrides[root]
            continue
        # 3. substring either direction
        hit = None
        for k, tab in prop_keys.items():
            if root in k or k in root:
                # Prefer the longest overlap to avoid "nav" matching too many
                if hit is None or abs(len(k) - len(root)) < abs(len(hit[0]) - len(root)):
                    hit = (k, tab)
        if hit:
            mapping[dom] = hit[1]
        else:
            unmatched.append(dom)
    return mapping, sorted(set(unmatched))


# ---------------------------------------------------------------------------
# Publisher Conversion Score (PCS) — secondary workbook
# ---------------------------------------------------------------------------
#
# The PCS report is a single sheet with two "Publisher Conversion Score"
# columns: the current period and a prior period for week-over-week reference.
# Column headers carry the date ranges (e.g. "Publisher Conversion Score April
# 13th-22nd" and "Publisher Conversion Score (March 30th-April 5th)") — we
# read those labels and flow them through to the UI so the dashboard stays
# in sync as the source is refreshed.


def _strip_publisher_prefix(desc: str) -> str:
    """Strip the 'Userwave - OEM Network N -' prefix from a publisher description."""
    if not isinstance(desc, str):
        return ""
    s = desc.strip()
    # e.g. "Userwave - OEM Network 4 - Daily Wire" -> "Daily Wire"
    s = re.sub(
        r"^userwave\s*[-–]\s*oem\s*network\s*\d+\s*[-–]?\s*",
        "",
        s,
        flags=re.IGNORECASE,
    )
    return s.strip(" -")


def _extract_period_label(header: str) -> str:
    """Pull the period text out of a PCS column header. Returns a clean label.

    'Publisher Conversion Score April 13th-22nd'            -> 'April 13th-22nd'
    'Publisher Conversion Score (March 30th-April 5th)'     -> 'March 30th-April 5th'
    """
    s = str(header).strip()
    s = re.sub(r"(?i)^publisher\s+conversion\s+score\s*", "", s)
    s = s.strip().strip("()").strip()
    return s or "PCS"


def parse_pcs_workbook(pcs_xlsx: Path) -> dict:
    """Parse the PCS comparison sheet. Returns dict with:
        tab_label, period_current, period_prior,
        rows=[{publisher, publisher_short, publisher_url, pcs_current, pcs_prior}]
    """
    wb = openpyxl.load_workbook(pcs_xlsx, data_only=True, read_only=False)
    if not wb.sheetnames:
        return {"tab_label": None, "period_current": None, "period_prior": None, "rows": []}

    # One tab per report; take the first populated one.
    ws = None
    for name in wb.sheetnames:
        sh = wb[name]
        if sh.max_row and sh.max_column:
            ws = sh
            break
    if ws is None:
        return {"tab_label": None, "period_current": None, "period_prior": None, "rows": []}

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"tab_label": ws.title, "period_current": None, "period_prior": None, "rows": []}

    # Resolve header row (row starting with something like 'Publisher - ID').
    header_idx = None
    for i, r in enumerate(rows[:5]):
        if r and isinstance(r[0], str) and "publisher" in r[0].strip().lower():
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    col = {h.lower(): i for i, h in enumerate(headers) if h}

    def find_col(*needles):
        for key, idx in col.items():
            low = key.lower()
            if all(n in low for n in needles):
                return idx
        return None

    i_desc = find_col("publisher", "description")
    i_url = find_col("publisher", "url")

    # All 'publisher conversion score' columns in left-to-right order.
    pcs_cols: list[tuple[int, str]] = []  # (col_index, raw_header)
    for idx, h in enumerate(headers):
        if h and "conversion score" in h.lower():
            pcs_cols.append((idx, h))

    if len(pcs_cols) < 1:
        print("[etl] warning: PCS sheet has no 'Publisher Conversion Score' column", file=sys.stderr)
        return {"tab_label": ws.title, "period_current": None, "period_prior": None, "rows": []}

    i_current, header_current = pcs_cols[0]
    i_prior, header_prior = (pcs_cols[1] if len(pcs_cols) >= 2 else (None, ""))

    period_current = _extract_period_label(header_current)
    period_prior = _extract_period_label(header_prior) if i_prior is not None else None

    out = []
    for r in rows[header_idx + 1:]:
        if not r or i_desc is None or i_desc >= len(r):
            continue
        desc = r[i_desc]
        if not isinstance(desc, str) or not desc.strip():
            continue
        url = r[i_url] if (i_url is not None and i_url < len(r)) else None
        cur = to_float(r[i_current]) if i_current < len(r) else None
        prv = (to_float(r[i_prior]) if (i_prior is not None and i_prior < len(r)) else None)
        out.append({
            "publisher": desc.strip(),
            "publisher_short": _strip_publisher_prefix(desc),
            "publisher_url": (url or "").strip() if isinstance(url, str) else url,
            "pcs_current": cur,
            "pcs_prior": prv,
        })
    return {
        "tab_label": ws.title,
        "period_current": period_current,
        "period_prior": period_prior,
        "rows": out,
    }


def build_pcs_property_map(
    property_tab_names: list[str],
    pcs_rows: list[dict],
) -> tuple[dict[str, dict], list[dict]]:
    """Join PCS rows to property tabs. Returns (map keyed by property_tab, unmatched rows)."""
    # Same approach as IVT: normalize both sides, exact key first, then overrides,
    # then substring match.
    overrides = {
        # publisher short name (normalized) -> property tab name
        "clickticks": "OEM5- Click Ticks ",
        "dailydash": "OEM4 -Daily Dash ",
        "starthubs": "OEM4 - Starthubs ",
        "startpage": "OEM4 - Startpage ",
        "spendider": "OEM5 - Spendider ",
        "swiftnavi": "OEM4 - Swiftnavi ",
        "promptnavi": "OEM4 - Promptnavi ",
        "easynav": "OEM5 - Easynav ",
        "naviseeking": "OEM4 - Naviseeking ",
        "navitravel": "OEM4 - Navitravel ",
    }
    prop_keys = {property_key(t): t for t in property_tab_names}

    mapping: dict[str, dict] = {}
    unmatched: list[dict] = []
    for row in pcs_rows:
        short = row.get("publisher_short") or row.get("publisher") or ""
        key = re.sub(r"[^a-z0-9]", "", short.lower())
        url = row.get("publisher_url") or ""
        url_key = domain_root(url) if isinstance(url, str) else ""

        match_tab = None
        # 1. exact short-name key
        if key and key in prop_keys:
            match_tab = prop_keys[key]
        # 2. exact URL root key
        elif url_key and url_key in prop_keys:
            match_tab = prop_keys[url_key]
        # 3. overrides (by short name)
        elif key in overrides and overrides[key] in property_tab_names:
            match_tab = overrides[key]
        elif url_key in overrides and overrides[url_key] in property_tab_names:
            match_tab = overrides[url_key]
        else:
            # 4. substring fuzzy
            for candidate in (key, url_key):
                if not candidate:
                    continue
                best = None
                for k, tab in prop_keys.items():
                    if candidate in k or k in candidate:
                        if best is None or abs(len(k) - len(candidate)) < abs(len(best[0]) - len(candidate)):
                            best = (k, tab)
                if best:
                    match_tab = best[1]
                    break

        if match_tab:
            # Keep first-seen (tab list isn't expected to have dupes, but be safe).
            mapping.setdefault(match_tab, row)
        else:
            unmatched.append(row)
    return mapping, unmatched


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build(source_xlsx: Path, output_html: Path, pcs_xlsx: Path | None = None) -> dict:
    wb = openpyxl.load_workbook(source_xlsx, data_only=True, read_only=False)

    revenue_rows: list[dict] = []
    property_meta: dict[str, dict] = {}

    for tab in wb.sheetnames:
        if tab in NON_PROPERTY_TABS:
            continue
        ws = wb[tab]
        rows = parse_property_tab(ws)
        if not rows:
            continue
        oem_group = parse_oem_group(tab)
        label = clean_property_label(tab)
        property_meta[tab] = {
            "tab": tab,
            "label": label,
            "oem_group": oem_group,
        }
        for r in rows:
            r["property_tab"] = tab
            r["property_label"] = label
            r["oem_group"] = oem_group
            revenue_rows.append(r)

    ivt_rows = parse_ivt_tab(wb["IVT Report"]) if "IVT Report" in wb.sheetnames else []

    domains = sorted({r["origin_domain"] for r in ivt_rows})
    domain_map, unmatched_domains = build_domain_property_map(
        list(property_meta.keys()), domains
    )
    for r in ivt_rows:
        r["matched_property_tab"] = domain_map.get(r["origin_domain"])

    # Compute high-level facts for the page header
    all_dates = sorted({r["date"] for r in revenue_rows})
    ivt_dates = sorted({r["date"] for r in ivt_rows})

    # ----- Performance tab: per-(property, day) records joined to IVT -----
    # Aggregate IVT to (matched_property_tab, date)
    ivt_pd: dict[tuple[str, str], dict] = {}
    for r in ivt_rows:
        t = r.get("matched_property_tab")
        if not t:
            continue
        k = (t, r["date"])
        d = ivt_pd.setdefault(k, {"impressions": 0.0, "fraud": 0.0})
        d["impressions"] += r["impressions"] or 0
        d["fraud"] += r["fraud"] or 0

    performance_pairs = []
    for r in revenue_rows:
        i = ivt_pd.get((r["property_tab"], r["date"]))
        if not i or not i["impressions"]:
            continue
        if not r.get("page_views") or r["page_views"] < 100:
            continue
        if not r.get("ad_revenue") or r["ad_revenue"] <= 0:
            continue
        rpm = (r["ad_revenue"] / r["page_views"]) * 1000
        fraud_pct = i["fraud"] / i["impressions"]
        performance_pairs.append({
            "date": r["date"],
            "property_tab": r["property_tab"],
            "property_label": r["property_label"],
            "oem_group": r["oem_group"],
            "page_views": r["page_views"],
            "ad_clicks": r.get("ad_clicks") or 0,
            "ad_revenue": r["ad_revenue"],
            "rpm": rpm,
            "impressions": i["impressions"],
            "fraud": i["fraud"],
            "fraud_pct": fraud_pct,
        })
    perf_dates = sorted({r["date"] for r in performance_pairs})

    # ----- PCS (Publisher Conversion Score) from secondary workbook -----
    pcs_payload: dict = {
        "tab_label": None, "period_current": None, "period_prior": None,
        "by_property": {}, "unmatched": [],
    }
    if pcs_xlsx is not None:
        try:
            pcs_raw = parse_pcs_workbook(pcs_xlsx)
            pcs_map, pcs_unmatched = build_pcs_property_map(
                list(property_meta.keys()), pcs_raw["rows"]
            )
            pcs_payload = {
                "tab_label": pcs_raw["tab_label"],
                "period_current": pcs_raw["period_current"],
                "period_prior": pcs_raw["period_prior"],
                "source_url": f"https://docs.google.com/spreadsheets/d/{PCS_SHEET_ID}/edit",
                "by_property": {
                    tab: {
                        "publisher": row.get("publisher"),
                        "publisher_short": row.get("publisher_short"),
                        "publisher_url": row.get("publisher_url"),
                        "pcs_current": row.get("pcs_current"),
                        "pcs_prior": row.get("pcs_prior"),
                    }
                    for tab, row in pcs_map.items()
                },
                "unmatched": [
                    {
                        "publisher": r.get("publisher"),
                        "publisher_short": r.get("publisher_short"),
                        "publisher_url": r.get("publisher_url"),
                        "pcs_current": r.get("pcs_current"),
                        "pcs_prior": r.get("pcs_prior"),
                    }
                    for r in pcs_unmatched
                ],
            }
            print(
                f"[etl] pcs: tab={pcs_raw['tab_label']!r} "
                f"periods=({pcs_raw['period_current']} | {pcs_raw['period_prior']}) "
                f"matched={len(pcs_map)} unmatched={len(pcs_unmatched)}",
                file=sys.stderr,
            )
        except Exception as e:
            print(f"[etl] warning: PCS parse failed: {e}", file=sys.stderr)

    payload = {
        "generated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source_url": f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit",
        "revenue": revenue_rows,
        "ivt": ivt_rows,
        "properties": list(property_meta.values()),
        "domain_map": domain_map,
        "unmatched_domains": unmatched_domains,
        "pcs": pcs_payload,
        "date_range": {
            "min": all_dates[0] if all_dates else None,
            "max": all_dates[-1] if all_dates else None,
            "ivt_min": ivt_dates[0] if ivt_dates else None,
            "ivt_max": ivt_dates[-1] if ivt_dates else None,
        },
        # Performance tab data (joined records for IVT × revenue analysis)
        "performance_pairs": performance_pairs,
        "performance_date_range": {
            "min": perf_dates[0] if perf_dates else None,
            "max": perf_dates[-1] if perf_dates else None,
        },
    }

    html = render_html(payload)
    output_html.write_text(html, encoding="utf-8")
    print(
        f"[etl] wrote {output_html} | "
        f"{len(revenue_rows)} revenue rows | {len(ivt_rows)} ivt rows | "
        f"{len(property_meta)} properties | {len(unmatched_domains)} unmatched domains",
        file=sys.stderr,
    )
    return payload


def render_html(payload: dict) -> str:
    """Render the dashboard. Data is embedded as JSON; UI runs in the browser."""
    data_json = json.dumps(payload, default=str)
    return HTML_TEMPLATE.replace("__DATA_JSON__", data_json)


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Kevin Dashboard — Revenue & IVT</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
  :root {
    --bg:#0f1115; --panel:#181b22; --panel2:#1f232c; --line:#2a2f3a;
    --text:#e8eaf0; --muted:#8b93a7; --accent:#6aa9ff; --good:#4ade80;
    --warn:#fbbf24; --bad:#f87171;
  }
  * { box-sizing: border-box; }
  html, body { margin:0; padding:0; background:var(--bg); color:var(--text);
    font:14px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }
  header { padding:20px 24px 0; border-bottom:1px solid var(--line);
    display:flex; justify-content:space-between; align-items:flex-end; gap:16px; flex-wrap:wrap;}
  header h1 { margin:0; font-size:20px; font-weight:600; letter-spacing:-0.01em;}
  header .meta { font-size:12px; color:var(--muted); text-align:right; }
  header .meta a { color:var(--accent); text-decoration:none; }
  /* Tab nav */
  .tab-nav { display:flex; gap:4px; margin-top:14px; flex:1 1 auto;}
  .tab-nav button { background:transparent; color:var(--muted); border:none;
    padding:10px 18px 12px; cursor:pointer; font:inherit; font-size:14px;
    border-bottom:2px solid transparent; letter-spacing:0.02em; }
  .tab-nav button:hover { color:var(--text); }
  .tab-nav button.active { color:var(--text); border-bottom-color:var(--accent); font-weight:600; }
  .tab-content[hidden] { display:none; }
  main { padding:20px 24px 40px; max-width:1600px; margin:0 auto; }
  .filters { display:grid; grid-template-columns: repeat(auto-fit, minmax(220px,1fr));
    gap:12px; background:var(--panel); padding:16px; border-radius:12px;
    border:1px solid var(--line); margin-bottom:20px;}
  .filter label { display:block; font-size:11px; text-transform:uppercase; letter-spacing:0.05em;
    color:var(--muted); margin-bottom:6px; }
  .filter select, .filter input {
    width:100%; background:var(--panel2); color:var(--text);
    border:1px solid var(--line); border-radius:8px; padding:8px 10px; font:inherit; }
  .preset-buttons { display:flex; gap:6px; flex-wrap:wrap; }
  .preset-buttons button { background:var(--panel2); color:var(--text); border:1px solid var(--line);
    border-radius:6px; padding:6px 10px; cursor:pointer; font:inherit; }
  .preset-buttons button.active { background:var(--accent); color:#0a1020; border-color:var(--accent); font-weight:600;}
  .kpis { display:grid; grid-template-columns:repeat(auto-fit, minmax(180px,1fr));
    gap:12px; margin-bottom:20px;}
  .kpi { background:var(--panel); border:1px solid var(--line); border-radius:12px; padding:14px 16px;}
  .kpi .label { font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:var(--muted);}
  .kpi .value { font-size:22px; font-weight:600; margin-top:6px;}
  .kpi .delta { font-size:12px; color:var(--muted); margin-top:4px;}
  .grid-2 { display:grid; grid-template-columns: 1fr 1fr; gap:16px; margin-bottom:20px;}
  @media (max-width: 1100px) { .grid-2 { grid-template-columns: 1fr; } }
  .card { background:var(--panel); border:1px solid var(--line); border-radius:12px; padding:14px 16px;}
  .card h3 { margin:0 0 12px; font-size:13px; text-transform:uppercase; letter-spacing:0.05em; color:var(--muted); font-weight:600;}
  .chart-wrap { position:relative; height:280px;}
  table { width:100%; border-collapse:collapse; font-size:13px; }
  th, td { text-align:left; padding:8px 10px; border-bottom:1px solid var(--line); white-space:nowrap;}
  th { font-size:11px; text-transform:uppercase; color:var(--muted); letter-spacing:0.05em;
    cursor:pointer; user-select:none; position:sticky; top:0; background:var(--panel);}
  th .sort { color:var(--muted); margin-left:4px; font-size:10px; }
  td.num, th.num { text-align:right; font-variant-numeric: tabular-nums; }
  tr:hover td { background:rgba(255,255,255,0.02); }
  /* Pivot matrix specific */
  table.matrix th.col-date { font-size:10px; padding:6px 6px; min-width:64px; }
  table.matrix td.cell { padding:4px 6px; font-size:12px; }
  table.matrix td.heat0 { background:transparent; }
  table.matrix td.heat1 { background:rgba(106,169,255,0.08); }
  table.matrix td.heat2 { background:rgba(106,169,255,0.18); }
  table.matrix td.heat3 { background:rgba(106,169,255,0.30); }
  table.matrix td.heat4 { background:rgba(106,169,255,0.45); color:#0a1020; font-weight:600; }
  table.matrix td.heat5 { background:rgba(106,169,255,0.65); color:#0a1020; font-weight:600; }
  table.matrix th.sticky, table.matrix td.sticky { position:sticky; left:0; background:var(--panel); z-index:1;}
  table.matrix th.sticky2, table.matrix td.sticky2 { position:sticky; background:var(--panel); z-index:1;}
  table.matrix th.sticky3, table.matrix td.sticky3 { position:sticky; background:var(--panel); z-index:1;}
  table.matrix th.sticky { left:0; min-width:60px;}
  table.matrix th.sticky2, table.matrix td.sticky2 { left:60px; min-width:200px;}
  table.matrix th.sticky3, table.matrix td.sticky3 { left:260px; min-width:110px; border-right:2px solid var(--line);}
  .pill { display:inline-block; padding:2px 8px; border-radius:999px; font-size:11px; font-weight:600;}
  .pill.good { background:rgba(74,222,128,0.15); color:var(--good); }
  .pill.warn { background:rgba(251,191,36,0.18); color:var(--warn); }
  .pill.bad  { background:rgba(248,113,113,0.18); color:var(--bad); }
  .pill.muted{ background:rgba(139,147,167,0.18); color:var(--muted); }
  .pill.orange{ background:rgba(251,146,60,0.18); color:#fb923c; }
  .pill.purple{ background:rgba(167,139,250,0.18); color:#a78bfa; }
  .scrollx { overflow-x:auto; max-height:520px; overflow-y:auto;}
  .multiselect { background:var(--panel2); border:1px solid var(--line); border-radius:8px; padding:6px;
    max-height:160px; overflow-y:auto; }
  .multiselect label { display:flex; align-items:center; gap:6px; padding:2px 4px; font-size:12px; color:var(--text); text-transform:none; letter-spacing:0; margin:0; cursor:pointer;}
  .multiselect label:hover { background:rgba(255,255,255,0.04); border-radius:4px;}
  .ms-empty { padding:8px 6px; font-size:12px; color:var(--muted); font-style:italic; text-align:center; }
  .ms-search { width:100%; background:var(--panel2); color:var(--text); border:1px solid var(--line);
    border-radius:6px; padding:5px 8px; font:inherit; font-size:12px; margin-bottom:6px;}
  .ms-search::placeholder { color:var(--muted); }
  .ms-meta { display:flex; justify-content:space-between; align-items:center; font-size:11px;
    color:var(--muted); text-transform:none; letter-spacing:0; margin:4px 2px 0; }
  .ms-actions { display:flex; gap:6px; margin-top:6px;}
  .ms-actions button { flex:1; background:var(--panel2); color:var(--muted); border:1px solid var(--line);
    border-radius:6px; padding:4px; cursor:pointer; font-size:11px;}
  details { background:var(--panel); border:1px solid var(--line); border-radius:12px; padding:10px 14px; margin-top:16px;}
  details summary { cursor:pointer; font-size:12px; color:var(--muted); text-transform:uppercase; letter-spacing:0.05em;}
  .empty { color:var(--muted); padding:20px; text-align:center; font-style:italic;}
</style>
</head>
<body>
<header>
  <div>
    <h1>Kevin Dashboard</h1>
    <div style="font-size:12px;color:var(--muted);margin-top:4px;">
      <span id="dateRangeLabel"></span>
    </div>
  </div>
  <nav class="tab-nav">
    <button class="tab-btn active" data-tab="overview">Overview</button>
    <button class="tab-btn" data-tab="performance">Performance</button>
  </nav>
  <div class="meta">
    <div>Generated <span id="gen"></span></div>
    <div><a id="srcLink" target="_blank">View source sheet</a></div>
  </div>
</header>

<main>
<div id="tab-overview" class="tab-content">
  <div class="filters">
    <div class="filter">
      <label>Date preset</label>
      <div class="preset-buttons" id="presets">
        <button data-d="yesterday">Yesterday</button>
        <button data-d="mtd" class="active">Month to date</button>
        <button data-d="7">Last 7d</button>
        <button data-d="14">Last 14d</button>
        <button data-d="30">Last 30d</button>
        <button data-d="90">Last 90d</button>
        <button data-d="all">All</button>
      </div>
    </div>
    <div class="filter">
      <label>From</label>
      <input type="date" id="from">
    </div>
    <div class="filter">
      <label>To</label>
      <input type="date" id="to">
    </div>
    <div class="filter">
      <label>OEM groups</label>
      <input type="search" class="ms-search" id="oemGroupsSearch" placeholder="Search OEM groups…" autocomplete="off">
      <div class="multiselect" id="oemGroups"></div>
      <div class="ms-meta"><span id="oemGroupsCount"></span></div>
      <div class="ms-actions">
        <button data-target="oemGroups" data-act="all">Select all</button>
        <button data-target="oemGroups" data-act="none">Clear</button>
      </div>
    </div>
    <div class="filter">
      <label>Properties</label>
      <input type="search" class="ms-search" id="propertiesSearch" placeholder="Search properties…" autocomplete="off">
      <div class="multiselect" id="properties"></div>
      <div class="ms-meta"><span id="propertiesCount"></span></div>
      <div class="ms-actions">
        <button data-target="properties" data-act="all">Select all</button>
        <button data-target="properties" data-act="none">Clear</button>
      </div>
    </div>
  </div>

  <div class="kpis" id="kpis"></div>

  <div class="grid-2">
    <div class="card">
      <h3>Daily revenue (filtered)</h3>
      <div class="chart-wrap"><canvas id="revChart"></canvas></div>
    </div>
    <div class="card">
      <h3>Network IVT — fraud % &amp; impressions</h3>
      <div class="chart-wrap"><canvas id="ivtChart"></canvas></div>
    </div>
  </div>

  <div class="card" style="margin-bottom:20px;">
    <h3>By date — daily summary</h3>
    <div class="scrollx">
      <table id="dateTable">
        <thead><tr>
          <th data-k="date">Date</th>
          <th class="num" data-k="ad_revenue">Ad Revenue</th>
          <th class="num" data-k="page_views">Page Views</th>
          <th class="num" data-k="ad_clicks">Ad Clicks</th>
          <th class="num" data-k="ctr">CTR</th>
          <th class="num" data-k="rpm">RPM</th>
          <th class="num" data-k="impressions">IVT Impr</th>
          <th class="num" data-k="fraud">IVT Fraud</th>
          <th class="num" data-k="fraud_pct">Fraud %</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <div class="card">
    <h3>
      By property — breakout (revenue + matched IVT)
      <span id="pcsWeekBadge" style="text-transform:none;letter-spacing:0;font-weight:400;font-size:12px;color:var(--muted);margin-left:10px;"></span>
    </h3>
    <div class="scrollx">
      <table id="propTable">
        <thead><tr>
          <th data-k="oem_group">OEM</th>
          <th data-k="label">Property</th>
          <th data-k="domain">Matched Domain</th>
          <th class="num" data-k="ad_revenue">Ad Revenue</th>
          <th class="num" data-k="page_views">Page Views</th>
          <th class="num" data-k="ad_clicks">Ad Clicks</th>
          <th class="num" data-k="cpc">CPC</th>
          <th class="num" data-k="ctr">CTR</th>
          <th class="num" data-k="rpm">RPM</th>
          <th class="num" data-k="vrpm">vRPM</th>
          <th class="num" data-k="impressions">IVT Impr</th>
          <th class="num" data-k="fraud_pct">Fraud %</th>
          <th class="num" data-k="pcs_current" id="pcsCurrentHeader">PCS current</th>
          <th class="num" data-k="pcs_prior" id="pcsPriorHeader">PCS prior</th>
          <th data-k="flag">Flag</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <div class="card" style="margin-top:20px;">
    <h3 style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap;">
      <span>Property × Day matrix</span>
      <span style="display:flex;gap:8px;align-items:center;text-transform:none;letter-spacing:0;font-weight:400;">
        <span style="color:var(--muted);font-size:12px;">Metric:</span>
        <select id="matrixMetric" style="background:var(--panel2);color:var(--text);border:1px solid var(--line);border-radius:6px;padding:4px 8px;font:inherit;">
          <option value="ad_revenue" selected>Ad Revenue</option>
          <option value="page_views">Page Views</option>
          <option value="ad_clicks">Ad Clicks</option>
          <option value="ctr">CTR</option>
          <option value="rpm">RPM</option>
          <option value="fraud_pct">Fraud %</option>
        </select>
        <label style="font-size:12px;color:var(--muted);display:flex;align-items:center;gap:4px;">
          <input type="checkbox" id="matrixHeatmap" checked> heatmap
        </label>
      </span>
    </h3>
    <div class="scrollx" style="max-height:600px;">
      <table id="matrixTable" class="matrix"><thead></thead><tbody></tbody></table>
    </div>
  </div>

  <div class="card" style="margin-top:20px;">
    <h3 style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap;">
      <span>Property daily detail (long form)</span>
      <span style="text-transform:none;letter-spacing:0;font-weight:400;font-size:12px;color:var(--muted);">
        Sortable; respects current filters. Showing top <span id="detailLimit">500</span> rows by date desc.
      </span>
    </h3>
    <div class="scrollx" style="max-height:520px;">
      <table id="detailTable">
        <thead><tr>
          <th data-k="date">Date</th>
          <th data-k="oem_group">OEM</th>
          <th data-k="property_label">Property</th>
          <th class="num" data-k="ad_revenue">Ad Revenue</th>
          <th class="num" data-k="page_views">Page Views</th>
          <th class="num" data-k="ad_clicks">Ad Clicks</th>
          <th class="num" data-k="ctr">CTR</th>
          <th class="num" data-k="ad_rpm">RPM</th>
          <th class="num" data-k="ad_vrpm">vRPM</th>
          <th class="num" data-k="impressions">IVT Impr</th>
          <th class="num" data-k="fraud_pct">Fraud %</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <details style="margin-top:16px;">
    <summary>IVT — origin domains not matched to a property tab (<span id="unmatchedCount">0</span>)</summary>
    <div id="unmatchedList" style="margin-top:8px;font-size:13px;color:var(--muted);"></div>
  </details>
</div><!-- /tab-overview -->

<div id="tab-performance" class="tab-content" hidden>
  <div class="card filter-card" style="padding:14px 22px;margin-bottom:18px;">
    <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
      <span style="font-size:11px;text-transform:uppercase;letter-spacing:0.05em;color:var(--muted);">Date range</span>
      <div class="preset-buttons" id="perfPresets">
        <button data-d="yesterday">Yesterday</button>
        <button data-d="7">Last 7d</button>
        <button data-d="14">Last 14d</button>
        <button data-d="30">Last 30d</button>
        <button data-d="90">Last 90d</button>
        <button data-d="all" class="active">All</button>
      </div>
      <div style="display:flex;gap:8px;align-items:center;">
        <label style="font-size:11px;color:var(--muted);text-transform:uppercase;">From</label>
        <input type="date" id="perfFrom" style="background:var(--panel2);color:var(--text);border:1px solid var(--line);border-radius:6px;padding:6px 8px;font:inherit;">
        <label style="font-size:11px;color:var(--muted);text-transform:uppercase;">To</label>
        <input type="date" id="perfTo" style="background:var(--panel2);color:var(--text);border:1px solid var(--line);border-radius:6px;padding:6px 8px;font:inherit;">
      </div>
      <span id="perfFilterMeta" style="margin-left:auto;font-size:12px;color:var(--muted);"></span>
    </div>
  </div>

  <div class="card" style="margin-bottom:18px;">
    <div class="verdict" id="perfVerdict" style="font-size:16px;line-height:1.55;"></div>
    <div id="perfLevels" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:14px;margin-top:16px;"></div>
    <div style="font-size:12px;color:var(--muted);margin-top:14px;">
      Pearson <em>r</em> ranges from −1 (perfectly inverse) to +1 (perfectly positive); <em>0</em> means no linear relationship.
      <em>p</em> below 0.05 is conventionally "statistically significant".
    </div>
  </div>

  <h2 style="font-size:15px;text-transform:uppercase;letter-spacing:0.06em;color:var(--muted);margin:30px 0 12px;font-weight:600;">Advertiser demand — what advertisers pay per click</h2>
  <div class="card" style="margin-bottom:18px;">
    <div style="font-size:13px;color:var(--muted);margin-bottom:12px;line-height:1.55;">
      <strong style="color:var(--text);">Why this view:</strong>
      RPM = CTR × CPC. A high-RPM property might just have a clicky audience.
      <strong style="color:var(--text);">CPC</strong> (revenue per click) is what advertisers are actually willing to pay for this audience —
      independent of how often users click. The clean-CPC axis discounts for fraud since advertisers will re-bid down on inventory they
      discover is invalid. Properties on the right with clean (green) inventory and headroom on the volume axis are the strongest scale
      candidates.
    </div>
    <div class="chart-wrap" style="height:440px;"><canvas id="perfDemandScatter"></canvas></div>
    <div style="display:flex;gap:20px;flex-wrap:wrap;align-items:center;margin-top:10px;font-size:12px;color:var(--muted);">
      <span>One bubble per property.</span>
      <span style="display:flex;align-items:center;gap:6px;">
        <span style="width:8px;height:8px;border-radius:50%;background:#94a3b8;display:inline-block;"></span>
        <span style="width:14px;height:14px;border-radius:50%;background:#94a3b8;display:inline-block;"></span>
        <span style="width:22px;height:22px;border-radius:50%;background:#94a3b8;display:inline-block;"></span>
        bubble size = revenue captured
      </span>
      <span style="display:flex;align-items:center;gap:6px;">
        <span style="width:14px;height:14px;border-radius:50%;background:#f87171;display:inline-block;"></span>
        <span style="width:14px;height:14px;border-radius:50%;background:#fbbf24;display:inline-block;"></span>
        <span style="width:14px;height:14px;border-radius:50%;background:#4ade80;display:inline-block;"></span>
        color = fraud (red 20%+, amber 5–20%, green &lt;5%)
      </span>
      <span><strong style="color:var(--good);">Right + green</strong> = strong advertiser demand on clean inventory.
            <strong style="color:var(--accent);">Right + green + tall</strong> = highest-leverage scale targets.</span>
    </div>
  </div>

  <h2 style="font-size:15px;text-transform:uppercase;letter-spacing:0.06em;color:var(--muted);margin:30px 0 12px;font-weight:600;">Scale candidates — premium demand with headroom</h2>
  <div class="card" style="margin-bottom:18px;">
    <div style="font-size:13px;color:var(--muted);margin-bottom:14px;line-height:1.55;">
      Properties matching all three criteria: <strong style="color:var(--text);">clean CPC in the top half</strong> (advertisers pay a premium),
      <strong style="color:var(--text);">page views below median</strong> (room to grow without saturating), and
      <strong style="color:var(--text);">fraud &lt; 10%</strong> (clean enough that scale won't get penalized).
      The "+~$X at median volume" estimate is what the property would earn if scaled to the network's median volume at its current CPC and CTR.
    </div>
    <div id="perfScaleGrid" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:12px;"></div>
  </div>

  <h2 style="font-size:15px;text-transform:uppercase;letter-spacing:0.06em;color:var(--muted);margin:30px 0 12px;font-weight:600;">Threshold — at what fraud % does RPM drop?</h2>
  <div class="card" style="margin-bottom:18px;">
    <div id="perfThresholdHeadline" style="font-size:15px;line-height:1.55;"></div>
    <div class="grid-2" style="margin-top:14px;">
      <div>
        <div class="chart-wrap" style="height:300px;"><canvas id="perfBinChart"></canvas></div>
        <div style="font-size:12px;color:var(--muted);margin-top:6px;">Volume-weighted RPM by fraud % bucket. Dashed line = lowest-fraud baseline.</div>
      </div>
      <div>
        <div class="chart-wrap" style="height:300px;"><canvas id="perfWithinBinChart"></canvas></div>
        <div style="font-size:12px;color:var(--muted);margin-top:6px;">Within-property: each day's RPM divided by its property's average. 1.0 = "at this property's typical RPM".</div>
      </div>
    </div>
    <div style="overflow:auto;margin-top:14px;max-height:520px;">
      <table id="perfBinTable" class="sortable">
        <thead><tr>
          <th data-k="bin_order">Fraud %</th>
          <th class="num" data-k="n">n property-days</th>
          <th class="num" data-k="weighted_rpm">Weighted RPM</th>
          <th class="num" data-k="mean_rpm">Mean RPM</th>
          <th class="num" data-k="median_rpm">Median RPM</th>
          <th class="num" data-k="pct_of_baseline">% of baseline</th>
          <th class="num" data-k="page_views">Page views</th>
          <th class="num" data-k="ad_revenue">Revenue</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <h2 style="font-size:15px;text-transform:uppercase;letter-spacing:0.06em;color:var(--muted);margin:30px 0 12px;font-weight:600;">Network daily — fraud % vs. RPM</h2>
  <div class="card grid-2" style="margin-bottom:18px;">
    <div>
      <div class="chart-wrap" style="height:300px;"><canvas id="perfNdScatter"></canvas></div>
      <div style="font-size:12px;color:var(--muted);margin-top:6px;">Each dot = one day across all matched properties.</div>
    </div>
    <div>
      <div class="chart-wrap" style="height:300px;"><canvas id="perfNdTime"></canvas></div>
      <div style="font-size:12px;color:var(--muted);margin-top:6px;">Same data over time — fraud % (left) and RPM (right).</div>
    </div>
  </div>

  <h2 style="font-size:15px;text-transform:uppercase;letter-spacing:0.06em;color:var(--muted);margin:30px 0 12px;font-weight:600;">Per-property — fraud × RPM × CTR × volume</h2>
  <div class="card" style="margin-bottom:18px;">
    <div class="chart-wrap" style="height:420px;"><canvas id="perfPpScatter"></canvas></div>
    <div style="display:flex;gap:20px;flex-wrap:wrap;align-items:center;margin-top:10px;font-size:12px;color:var(--muted);">
      <span>One bubble per property. <strong>Top-left + green + large</strong> = candidates to scale up.</span>
      <span>Bubble size = page-view volume. Color = CTR (low → high tertile).</span>
    </div>
  </div>

  <div class="card" style="margin-bottom:18px;">
    <div style="font-size:12px;color:var(--muted);margin-bottom:10px;line-height:1.5;">
      <strong style="color:var(--text);">Demand tier</strong> ranks properties by clean CPC = CPC × (1 − fraud %) — the price advertisers
      effectively pay once invalid traffic is discounted. Tertiles within the current filter window:
      <span class="pill pos" style="font-size:10px;">High</span> top third &nbsp;·&nbsp;
      <span class="pill" style="background:rgba(251,191,36,0.18);color:var(--warn);font-size:10px;">Mid</span> middle third &nbsp;·&nbsp;
      <span class="pill" style="background:rgba(139,147,167,0.18);color:var(--muted);font-size:10px;">Low</span> bottom third.
      Hover any tier pill to see the exact CPC math.
    </div>
    <div style="overflow:auto;max-height:520px;">
      <table id="perfPropTable" class="sortable">
        <thead><tr>
          <th data-k="oem_group">OEM</th>
          <th data-k="label">Property</th>
          <th class="num" data-k="days">Days</th>
          <th class="num" data-k="cpc">CPC</th>
          <th class="num" data-k="cpc_clean">Clean CPC</th>
          <th class="num" data-k="rpm">RPM</th>
          <th class="num" data-k="ctr">CTR</th>
          <th class="num" data-k="fraud_pct">Fraud %</th>
          <th class="num" data-k="page_views">Page Views</th>
          <th class="num" data-k="ad_revenue">Revenue</th>
          <th class="num" data-k="focus_score">Score</th>
          <th data-k="demand_tier">Demand tier</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <h2 style="font-size:15px;text-transform:uppercase;letter-spacing:0.06em;color:var(--muted);margin:30px 0 12px;font-weight:600;">Per-property — within-property correlation strength</h2>
  <div class="card" style="margin-bottom:18px;">
    <p style="margin:0 0 10px;color:var(--muted);font-size:13px;">
      For each property with enough days, the correlation between that property's daily fraud % and its daily RPM.
      A strong negative <em>r</em> means: at this property, days with more fraud also tend to be days with lower RPM.
    </p>
    <div style="overflow:auto;max-height:520px;">
      <table id="perfPpCorrTable" class="sortable">
        <thead><tr>
          <th data-k="oem_group">OEM</th>
          <th data-k="label">Property</th>
          <th class="num" data-k="n">n days</th>
          <th class="num" data-k="r">r (fraud %, RPM)</th>
          <th class="num" data-k="mean_rpm">mean RPM</th>
          <th class="num" data-k="mean_fraud">mean fraud %</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>
</div><!-- /tab-performance -->
</main>

<script>
const DATA = __DATA_JSON__;

// --- helpers -----------------------------------------------------------
const fmt = {
  money: v => v == null ? "—" : "$" + v.toLocaleString(undefined,{maximumFractionDigits:2, minimumFractionDigits:2}),
  int: v => v == null ? "—" : Math.round(v).toLocaleString(),
  num: (v, d=2) => v == null ? "—" : v.toLocaleString(undefined,{maximumFractionDigits:d, minimumFractionDigits:d}),
  pct: v => v == null ? "—" : (v*100).toLocaleString(undefined,{maximumFractionDigits:2}) + "%",
  // IVT fraud_pct in the source is already a fraction (0..1)
  pctRaw: v => v == null ? "—" : (v*100).toLocaleString(undefined,{maximumFractionDigits:2}) + "%",
};
function sum(arr, k) { return arr.reduce((a,r)=>a+(r[k]||0),0); }
function mean(arr, k) {
  const v = arr.map(r=>r[k]).filter(x=>x!=null && !Number.isNaN(x));
  return v.length ? v.reduce((a,b)=>a+b,0)/v.length : null;
}

// --- header -------------------------------------------------------------
document.getElementById("gen").textContent = DATA.generated_at;
document.getElementById("srcLink").href = DATA.source_url;
document.getElementById("dateRangeLabel").textContent =
  `Source data: ${DATA.date_range.min ?? "?"} → ${DATA.date_range.max ?? "?"}  ·  IVT: ${DATA.date_range.ivt_min ?? "?"} → ${DATA.date_range.ivt_max ?? "?"}`;

// --- filter state -------------------------------------------------------
const state = {
  from: null, to: null,
  oemGroups: new Set(),
  properties: new Set(),
};

const allOemGroups = [...new Set(DATA.properties.map(p=>p.oem_group))].sort();
const allProperties = DATA.properties.slice().sort((a,b)=>a.label.localeCompare(b.label));
allOemGroups.forEach(g=>state.oemGroups.add(g));
allProperties.forEach(p=>state.properties.add(p.tab));

// Render multiselects with search filtering and a live selection counter.
// Select all / Clear operate on the VISIBLE (currently filtered) subset so you
// can search a substring and bulk-toggle just those matches.
const msQuery = { oemGroups: "", properties: "" };
function visibleItems(id, items, getLabel) {
  const q = (msQuery[id] || "").trim().toLowerCase();
  if (!q) return items;
  return items.filter(it => getLabel(it).toLowerCase().includes(q));
}
function renderMultiselect(id, items, getValue, getLabel, set) {
  const root = document.getElementById(id);
  root.innerHTML = "";
  const shown = visibleItems(id, items, getLabel);
  if (!shown.length) {
    const e = document.createElement("div");
    e.className = "ms-empty";
    e.textContent = "No matches";
    root.appendChild(e);
  } else {
    shown.forEach(it => {
      const v = getValue(it);
      const lbl = document.createElement("label");
      lbl.innerHTML = `<input type="checkbox" ${set.has(v)?"checked":""} value="${v}"> <span>${getLabel(it)}</span>`;
      lbl.querySelector("input").addEventListener("change", e => {
        if (e.target.checked) set.add(v); else set.delete(v);
        updateMsCount(id, items, getValue);
        render();
      });
      root.appendChild(lbl);
    });
  }
  updateMsCount(id, items, getValue);
}
function updateMsCount(id, items, getValue) {
  const el = document.getElementById(id + "Count");
  if (!el) return;
  const setObj = (id === "oemGroups") ? state.oemGroups : state.properties;
  const total = items.length;
  const selected = items.filter(it => setObj.has(getValue(it))).length;
  const q = (msQuery[id] || "").trim();
  if (q) {
    const shownCount = visibleItems(id, items, (id === "oemGroups") ? (g=>g) : (p=>`${p.label} (${p.oem_group})`)).length;
    el.textContent = `${selected} of ${total} selected · ${shownCount} shown`;
  } else {
    el.textContent = `${selected} of ${total} selected`;
  }
}
function repaintMultiselects() {
  renderMultiselect("oemGroups", allOemGroups, g=>g, g=>g, state.oemGroups);
  renderMultiselect("properties", allProperties, p=>p.tab, p=>`${p.label} (${p.oem_group})`, state.properties);
}
repaintMultiselects();

// Search input handlers
document.getElementById("oemGroupsSearch").addEventListener("input", e => {
  msQuery.oemGroups = e.target.value;
  renderMultiselect("oemGroups", allOemGroups, g=>g, g=>g, state.oemGroups);
});
document.getElementById("propertiesSearch").addEventListener("input", e => {
  msQuery.properties = e.target.value;
  renderMultiselect("properties", allProperties, p=>p.tab, p=>`${p.label} (${p.oem_group})`, state.properties);
});

document.querySelectorAll('.ms-actions button').forEach(b=>{
  b.addEventListener("click", ()=>{
    const target = b.dataset.target;
    const act = b.dataset.act;
    // Bulk actions apply to currently-visible items only so the user can
    // search for a substring and toggle that subset without disturbing the rest.
    if (target==="oemGroups") {
      const vis = visibleItems("oemGroups", allOemGroups, g=>g);
      vis.forEach(g => act==="all" ? state.oemGroups.add(g) : state.oemGroups.delete(g));
    } else {
      const vis = visibleItems("properties", allProperties, p=>`${p.label} (${p.oem_group})`);
      vis.forEach(p => act==="all" ? state.properties.add(p.tab) : state.properties.delete(p.tab));
    }
    repaintMultiselects();
    render();
  });
});

// --- date presets -------------------------------------------------------
function applyPreset(d) {
  document.querySelectorAll("#presets button").forEach(b=>b.classList.toggle("active", b.dataset.d===String(d)));
  const max = DATA.date_range.max ? new Date(DATA.date_range.max) : new Date();
  const min = DATA.date_range.min ? new Date(DATA.date_range.min) : new Date();
  let from = min, to = max;
  const key = String(d);
  if (key === "all") {
    // from=min, to=max (already set above)
  } else if (key === "yesterday") {
    // Calendar yesterday, clamped to the available data window.
    const y = new Date();
    y.setDate(y.getDate() - 1);
    if (y > max) { from = to = max; }
    else if (y < min) { from = to = min; }
    else { from = to = y; }
  } else if (key === "mtd") {
    // First of the current month (client local) through the latest data.
    const now = new Date();
    const first = new Date(now.getFullYear(), now.getMonth(), 1);
    from = first < min ? min : first;
    to = max;
  } else {
    const days = parseInt(key, 10);
    from = new Date(max);
    from.setDate(from.getDate() - (days - 1));
    if (from < min) from = min;
  }
  state.from = from.toISOString().slice(0,10);
  state.to = to.toISOString().slice(0,10);
  document.getElementById("from").value = state.from;
  document.getElementById("to").value = state.to;
  render();
}
document.querySelectorAll("#presets button").forEach(b=>{
  b.addEventListener("click", ()=>applyPreset(b.dataset.d));
});
document.getElementById("from").addEventListener("change", e=>{
  state.from = e.target.value;
  document.querySelectorAll("#presets button").forEach(b=>b.classList.remove("active"));
  render();
});
document.getElementById("to").addEventListener("change", e=>{
  state.to = e.target.value;
  document.querySelectorAll("#presets button").forEach(b=>b.classList.remove("active"));
  render();
});

// --- charts -------------------------------------------------------------
let revChart, ivtChart;
function chartColors() {
  return { line: "#6aa9ff", line2: "#fbbf24", bad: "#f87171", grid: "rgba(255,255,255,0.05)", text:"#8b93a7"};
}
function ensureCharts() {
  const c = chartColors();
  Chart.defaults.color = c.text;
  Chart.defaults.borderColor = c.grid;
  if (!revChart) {
    revChart = new Chart(document.getElementById("revChart"), {
      type:"line",
      data:{ labels:[], datasets:[{label:"Ad Revenue", data:[], borderColor:c.line, backgroundColor:"rgba(106,169,255,0.15)", tension:0.25, fill:true}] },
      options:{ responsive:true, maintainAspectRatio:false,
        scales:{y:{ticks:{callback:v=>"$"+v.toLocaleString()}}},
        plugins:{legend:{display:false}}
      }});
  }
  if (!ivtChart) {
    ivtChart = new Chart(document.getElementById("ivtChart"), {
      type:"line",
      data:{ labels:[], datasets:[
        {label:"Fraud %", data:[], borderColor:c.bad, backgroundColor:"rgba(248,113,113,0.15)", yAxisID:"y", tension:0.25, fill:true},
        {label:"Impressions", data:[], borderColor:c.line2, backgroundColor:"rgba(251,191,36,0.05)", yAxisID:"y1", tension:0.25, fill:false},
      ]},
      options:{ responsive:true, maintainAspectRatio:false,
        scales:{
          y:{position:"left", ticks:{callback:v=>(v*100).toFixed(0)+"%"}, title:{display:true, text:"Fraud %"}},
          y1:{position:"right", grid:{drawOnChartArea:false}, ticks:{callback:v=>v.toLocaleString()}, title:{display:true, text:"Impressions"}}
        },
        plugins:{legend:{labels:{boxWidth:12}}}
      }});
  }
}

// --- table sort state ---------------------------------------------------
const sortState = {
  dateTable:{k:"date", dir:-1},
  propTable:{k:"ad_revenue", dir:-1},
  detailTable:{k:"date", dir:-1},
  matrixTable:{k:"_total", dir:-1},
};
const matrixState = { metric: "ad_revenue", heatmap: true };
document.getElementById("matrixMetric").addEventListener("change", e=>{
  matrixState.metric = e.target.value; render();
});
document.getElementById("matrixHeatmap").addEventListener("change", e=>{
  matrixState.heatmap = e.target.checked; render();
});
function bindSort(tableId) {
  document.querySelectorAll(`#${tableId} th`).forEach(th=>{
    th.addEventListener("click", ()=>{
      const k = th.dataset.k;
      const s = sortState[tableId];
      if (s.k===k) s.dir = -s.dir; else { s.k = k; s.dir = -1; }
      render();
    });
  });
}
bindSort("dateTable");
bindSort("propTable");
bindSort("detailTable");
function sortRows(rows, s) {
  return rows.slice().sort((a,b)=>{
    const x = a[s.k], y = b[s.k];
    if (x==null && y==null) return 0;
    if (x==null) return 1;
    if (y==null) return -1;
    if (typeof x === "number") return (x-y)*s.dir;
    return String(x).localeCompare(String(y))*s.dir;
  });
}

// --- main render --------------------------------------------------------
function inDateRange(d) {
  return (!state.from || d>=state.from) && (!state.to || d<=state.to);
}
function inGroup(g) { return state.oemGroups.has(g); }
function inProperty(t) { return state.properties.has(t); }

function render() {
  ensureCharts();

  // filter revenue
  const rev = DATA.revenue.filter(r=>inDateRange(r.date) && inGroup(r.oem_group) && inProperty(r.property_tab));
  const ivt = DATA.ivt.filter(r=>inDateRange(r.date));
  // Per-property IVT (only matched + selected properties)
  const ivtMatched = ivt.filter(r=>r.matched_property_tab && state.properties.has(r.matched_property_tab));

  // KPIs
  const totalRev = sum(rev, "ad_revenue");
  const totalPv = sum(rev, "page_views");
  const totalClicks = sum(rev, "ad_clicks");
  const ctr = totalPv ? totalClicks/totalPv : null;
  const rpm = totalPv ? (totalRev/totalPv)*1000 : null;
  const totalImpr = sum(ivt, "impressions");
  const totalFraud = sum(ivt, "fraud");
  const networkFraudPct = totalImpr ? totalFraud/totalImpr : null;

  const kpis = [
    {label:"Total Ad Revenue", value: fmt.money(totalRev)},
    {label:"Page Views", value: fmt.int(totalPv)},
    {label:"Ad Clicks", value: fmt.int(totalClicks)},
    {label:"CTR", value: fmt.pct(ctr)},
    {label:"RPM", value: fmt.num(rpm,2)},
    {label:"IVT Impressions", value: fmt.int(totalImpr)},
    {label:"Network Fraud %", value: fmt.pctRaw(networkFraudPct), klass: networkFraudPct>0.2 ? "bad" : networkFraudPct>0.05 ? "warn" : "good"},
  ];
  document.getElementById("kpis").innerHTML = kpis.map(k=>{
    const cls = k.klass ? `style="color:var(--${k.klass==='bad'?'bad':k.klass==='warn'?'warn':'good'})"` : "";
    return `<div class="kpi"><div class="label">${k.label}</div><div class="value" ${cls}>${k.value}</div></div>`;
  }).join("");

  // Daily aggregations for table + charts
  const byDate = {};
  rev.forEach(r=>{
    const k = r.date;
    if (!byDate[k]) byDate[k] = {date:k, ad_revenue:0, page_views:0, ad_clicks:0};
    byDate[k].ad_revenue += r.ad_revenue||0;
    byDate[k].page_views += r.page_views||0;
    byDate[k].ad_clicks += r.ad_clicks||0;
  });
  const ivtByDate = {};
  ivt.forEach(r=>{
    const k = r.date;
    if (!ivtByDate[k]) ivtByDate[k] = {impressions:0, fraud:0, clicks:0};
    ivtByDate[k].impressions += r.impressions||0;
    ivtByDate[k].fraud += r.fraud||0;
    ivtByDate[k].clicks += r.clicks||0;
  });
  const allDates = [...new Set([...Object.keys(byDate), ...Object.keys(ivtByDate)])].sort();
  const dateRows = allDates.map(d=>{
    const a = byDate[d] || {ad_revenue:0, page_views:0, ad_clicks:0};
    const i = ivtByDate[d] || {impressions:0, fraud:0};
    return {
      date: d,
      ad_revenue: a.ad_revenue,
      page_views: a.page_views,
      ad_clicks: a.ad_clicks,
      ctr: a.page_views ? a.ad_clicks/a.page_views : null,
      rpm: a.page_views ? (a.ad_revenue/a.page_views)*1000 : null,
      impressions: i.impressions,
      fraud: i.fraud,
      fraud_pct: i.impressions ? i.fraud/i.impressions : null,
    };
  });
  const dateRowsSorted = sortRows(dateRows, sortState.dateTable);
  document.querySelector("#dateTable tbody").innerHTML = dateRowsSorted.length ? dateRowsSorted.map(r=>`
    <tr>
      <td>${r.date}</td>
      <td class="num">${fmt.money(r.ad_revenue)}</td>
      <td class="num">${fmt.int(r.page_views)}</td>
      <td class="num">${fmt.int(r.ad_clicks)}</td>
      <td class="num">${fmt.pct(r.ctr)}</td>
      <td class="num">${fmt.num(r.rpm,2)}</td>
      <td class="num">${fmt.int(r.impressions)}</td>
      <td class="num">${fmt.int(r.fraud)}</td>
      <td class="num">${pillFraud(r.fraud_pct)}</td>
    </tr>`).join("") : `<tr><td colspan="9" class="empty">No data in selected range.</td></tr>`;

  // Revenue chart: multi-line when ≤10 properties selected, otherwise aggregate
  const chartDates = dateRowsSorted.slice().sort((a,b)=>a.date.localeCompare(b.date));
  const chartLabels = chartDates.map(r=>r.date);
  const palette = ["#6aa9ff","#fbbf24","#4ade80","#f87171","#c084fc","#22d3ee","#fb923c","#a3e635","#f472b6","#94a3b8"];
  if (state.properties.size <= 10 && state.properties.size > 0) {
    // Build per-property daily revenue series
    const series = {};
    rev.forEach(r=>{
      if (!series[r.property_tab]) series[r.property_tab] = {label:r.property_label, data:{}};
      series[r.property_tab].data[r.date] = (series[r.property_tab].data[r.date]||0) + (r.ad_revenue||0);
    });
    const seriesArr = Object.entries(series).sort((a,b)=>a[1].label.localeCompare(b[1].label));
    revChart.data.labels = chartLabels;
    revChart.data.datasets = seriesArr.map(([tab, s], i)=>({
      label: s.label,
      data: chartLabels.map(d=>s.data[d]||0),
      borderColor: palette[i % palette.length],
      backgroundColor: "transparent",
      tension: 0.25,
      fill: false,
      borderWidth: 2,
      pointRadius: 1.5,
    }));
    revChart.options.plugins.legend.display = true;
    revChart.options.plugins.legend.position = "bottom";
  } else {
    revChart.data.labels = chartLabels;
    revChart.data.datasets = [{
      label:"Ad Revenue (all selected)",
      data: chartDates.map(r=>r.ad_revenue),
      borderColor:"#6aa9ff",
      backgroundColor:"rgba(106,169,255,0.15)",
      tension:0.25, fill:true,
    }];
    revChart.options.plugins.legend.display = false;
  }
  revChart.update();
  ivtChart.data.labels = chartDates.map(r=>r.date);
  ivtChart.data.datasets[0].data = chartDates.map(r=>r.fraud_pct||0);
  ivtChart.data.datasets[1].data = chartDates.map(r=>r.impressions);
  ivtChart.update();

  // Property breakout
  const byProp = {};
  rev.forEach(r=>{
    const k = r.property_tab;
    if (!byProp[k]) byProp[k] = {tab:k, label:r.property_label, oem_group:r.oem_group,
      ad_revenue:0, page_views:0, ad_clicks:0, vwa:0};
    byProp[k].ad_revenue += r.ad_revenue||0;
    byProp[k].page_views += r.page_views||0;
    byProp[k].ad_clicks += r.ad_clicks||0;
  });
  // Build IVT-by-property aggregates
  const ivtByProp = {};
  const domainByProp = {};
  ivtMatched.forEach(r=>{
    const k = r.matched_property_tab;
    if (!ivtByProp[k]) ivtByProp[k] = {impressions:0, fraud:0};
    ivtByProp[k].impressions += r.impressions||0;
    ivtByProp[k].fraud += r.fraud||0;
    if (!domainByProp[k]) domainByProp[k] = new Set();
    domainByProp[k].add(r.origin_domain);
  });
  const pcsByProp = (DATA.pcs && DATA.pcs.by_property) || {};
  const propRows = Object.values(byProp).map(p=>{
    const i = ivtByProp[p.tab] || {impressions:0, fraud:0};
    const fp = i.impressions ? i.fraud/i.impressions : null;
    const pcs = pcsByProp[p.tab] || {};
    return {
      ...p,
      ctr: p.page_views ? p.ad_clicks/p.page_views : null,
      cpc: p.ad_clicks ? p.ad_revenue/p.ad_clicks : null,
      rpm: p.page_views ? (p.ad_revenue/p.page_views)*1000 : null,
      vrpm: null,
      impressions: i.impressions,
      fraud_pct: fp,
      pcs_current: (pcs.pcs_current==null ? null : pcs.pcs_current),
      pcs_prior:   (pcs.pcs_prior==null   ? null : pcs.pcs_prior),
      domain: domainByProp[p.tab] ? [...domainByProp[p.tab]].join(", ") : "—",
      flag: fp==null ? "—" : (fp>0.2 ? "bad" : fp>0.05 ? "warn" : "good"),
    };
  });
  const propSorted = sortRows(propRows, sortState.propTable);
  document.querySelector("#propTable tbody").innerHTML = propSorted.length ? propSorted.map(r=>`
    <tr>
      <td><span class="pill muted">${r.oem_group}</span></td>
      <td>${r.label}</td>
      <td style="color:var(--muted);font-size:12px;">${r.domain}</td>
      <td class="num">${fmt.money(r.ad_revenue)}</td>
      <td class="num">${fmt.int(r.page_views)}</td>
      <td class="num">${fmt.int(r.ad_clicks)}</td>
      <td class="num">${r.cpc!=null?"$"+r.cpc.toFixed(3):"—"}</td>
      <td class="num">${fmt.pct(r.ctr)}</td>
      <td class="num">${fmt.num(r.rpm,2)}</td>
      <td class="num">—</td>
      <td class="num">${fmt.int(r.impressions)}</td>
      <td class="num">${pillFraud(r.fraud_pct)}</td>
      <td class="num">${pcsCell(r.pcs_current)}</td>
      <td class="num">${pcsCell(r.pcs_prior)}</td>
      <td>${flagPill(r.flag)}</td>
    </tr>`).join("") : `<tr><td colspan="15" class="empty">No properties in selected filters.</td></tr>`;

  // ---------- Property × Day matrix ----------
  renderMatrix(rev, ivtMatched);

  // ---------- Property daily detail (long form) ----------
  renderDetail(rev, ivtMatched);

  // Unmatched domains
  document.getElementById("unmatchedCount").textContent = DATA.unmatched_domains.length;
  document.getElementById("unmatchedList").innerHTML = DATA.unmatched_domains.length
    ? DATA.unmatched_domains.map(d=>`<span style="display:inline-block;background:var(--panel2);padding:3px 8px;border-radius:6px;margin:2px;">${d}</span>`).join("")
    : "<i>All IVT origin domains were matched to a property tab.</i>";
}
function renderMatrix(rev, ivtMatched) {
  const metric = matrixState.metric;
  const heat = matrixState.heatmap;
  // Need full date list (filtered)
  const dates = [...new Set(rev.map(r=>r.date))].sort();
  // Property index (only those with any data in the filtered window)
  const propIdx = {};
  rev.forEach(r=>{
    if (!propIdx[r.property_tab]) propIdx[r.property_tab] = {
      tab:r.property_tab, label:r.property_label, oem_group:r.oem_group,
      cells:{}, // date -> {ad_revenue, page_views, ad_clicks}
    };
    const c = propIdx[r.property_tab].cells[r.date] || (propIdx[r.property_tab].cells[r.date] = {ad_revenue:0, page_views:0, ad_clicks:0});
    c.ad_revenue += r.ad_revenue||0;
    c.page_views += r.page_views||0;
    c.ad_clicks += r.ad_clicks||0;
  });
  // For fraud_pct, fold in matched IVT per property+date
  const ivtCells = {}; // tab -> date -> {impressions, fraud}
  ivtMatched.forEach(r=>{
    const t = r.matched_property_tab;
    if (!ivtCells[t]) ivtCells[t] = {};
    const c = ivtCells[t][r.date] || (ivtCells[t][r.date] = {impressions:0, fraud:0});
    c.impressions += r.impressions||0;
    c.fraud += r.fraud||0;
  });

  // Compute per-cell metric value + per-row total
  function cellValue(prop, date) {
    const c = prop.cells[date];
    const i = (ivtCells[prop.tab] || {})[date];
    if (metric === "ad_revenue") return c ? c.ad_revenue : null;
    if (metric === "page_views") return c ? c.page_views : null;
    if (metric === "ad_clicks") return c ? c.ad_clicks : null;
    if (metric === "ctr") return (c && c.page_views) ? c.ad_clicks/c.page_views : null;
    if (metric === "rpm") return (c && c.page_views) ? (c.ad_revenue/c.page_views)*1000 : null;
    if (metric === "fraud_pct") return (i && i.impressions) ? i.fraud/i.impressions : null;
    return null;
  }
  function rowTotal(prop) {
    if (metric === "ad_revenue" || metric === "page_views" || metric === "ad_clicks") {
      return dates.reduce((a,d)=>a+(cellValue(prop,d)||0),0);
    }
    if (metric === "ctr" || metric === "rpm") {
      let pv=0, num=0;
      dates.forEach(d=>{
        const c = prop.cells[d]; if (!c) return;
        pv += c.page_views||0;
        num += metric==="ctr" ? (c.ad_clicks||0) : (c.ad_revenue||0);
      });
      return pv ? (metric==="ctr" ? num/pv : (num/pv)*1000) : null;
    }
    if (metric === "fraud_pct") {
      const ic = ivtCells[prop.tab]; if (!ic) return null;
      let imp=0, fr=0;
      dates.forEach(d=>{ const c=ic[d]; if (c) { imp+=c.impressions; fr+=c.fraud; }});
      return imp ? fr/imp : null;
    }
    return null;
  }

  // Sort properties by total or label
  const props = Object.values(propIdx).map(p => ({...p, _total: rowTotal(p)}));
  const s = sortState.matrixTable;
  if (s.k === "_total") {
    props.sort((a,b)=>{
      const x = a._total, y = b._total;
      if (x==null && y==null) return 0;
      if (x==null) return 1;
      if (y==null) return -1;
      return (x-y)*s.dir;
    });
  } else if (s.k === "label") {
    props.sort((a,b)=>a.label.localeCompare(b.label)*s.dir);
  } else if (s.k === "oem_group") {
    props.sort((a,b)=>a.oem_group.localeCompare(b.oem_group)*s.dir);
  }

  // Compute heatmap quintiles for non-null values across all cells
  let breaks = null;
  if (heat) {
    const allVals = [];
    props.forEach(p => dates.forEach(d => {
      const v = cellValue(p,d); if (v != null && v > 0) allVals.push(v);
    }));
    if (allVals.length) {
      allVals.sort((a,b)=>a-b);
      breaks = [0.2,0.4,0.6,0.8,0.95].map(q => allVals[Math.min(allVals.length-1, Math.floor(allVals.length*q))]);
    }
  }
  function heatClass(v) {
    if (!heat || v==null || v<=0 || !breaks) return "heat0";
    if (v <= breaks[0]) return "heat0";
    if (v <= breaks[1]) return "heat1";
    if (v <= breaks[2]) return "heat2";
    if (v <= breaks[3]) return "heat3";
    if (v <= breaks[4]) return "heat4";
    return "heat5";
  }
  function fmtCell(v) {
    if (v==null) return '<span style="color:var(--muted)">—</span>';
    if (metric==="ad_revenue") return "$"+v.toLocaleString(undefined,{maximumFractionDigits:0});
    if (metric==="page_views" || metric==="ad_clicks") return Math.round(v).toLocaleString();
    if (metric==="ctr" || metric==="fraud_pct") return (v*100).toFixed(2)+"%";
    if (metric==="rpm") return v.toFixed(2);
    return v;
  }
  function fmtTotal(v) {
    if (v==null) return "—";
    if (metric==="ad_revenue") return "$"+v.toLocaleString(undefined,{maximumFractionDigits:2});
    if (metric==="page_views" || metric==="ad_clicks") return Math.round(v).toLocaleString();
    if (metric==="ctr" || metric==="fraud_pct") return (v*100).toFixed(2)+"%";
    if (metric==="rpm") return v.toFixed(2);
    return v;
  }

  // Build table
  const head = `<tr>
    <th class="sticky" data-k="oem_group">OEM</th>
    <th class="sticky2" data-k="label">Property</th>
    <th class="sticky3 num" data-k="_total">Total</th>
    ${dates.map(d=>`<th class="num col-date">${d.slice(5)}</th>`).join("")}
  </tr>`;
  const body = props.map(p=>`
    <tr>
      <td class="sticky"><span class="pill muted">${p.oem_group}</span></td>
      <td class="sticky2">${p.label}</td>
      <td class="sticky3 num" style="border-right:2px solid var(--line);font-weight:600;">${fmtTotal(p._total)}</td>
      ${dates.map(d => {
        const v = cellValue(p,d);
        return `<td class="cell num ${heatClass(v)}">${fmtCell(v)}</td>`;
      }).join("")}
    </tr>`).join("");
  const tbl = document.getElementById("matrixTable");
  tbl.querySelector("thead").innerHTML = head;
  tbl.querySelector("tbody").innerHTML = props.length ? body :
    `<tr><td colspan="${dates.length+3}" class="empty">No data in selected filters.</td></tr>`;
  // Wire header sort (only the first 3 sortable columns)
  tbl.querySelectorAll("thead th[data-k]").forEach(th=>{
    th.addEventListener("click", ()=>{
      const k = th.dataset.k;
      if (s.k===k) s.dir = -s.dir; else { s.k = k; s.dir = -1; }
      render();
    });
  });
}

function renderDetail(rev, ivtMatched) {
  // Build long-form rows: one per (date, property), with matched IVT joined
  const ivtKey = {};
  ivtMatched.forEach(r=>{
    const k = r.date + "|" + r.matched_property_tab;
    if (!ivtKey[k]) ivtKey[k] = {impressions:0, fraud:0};
    ivtKey[k].impressions += r.impressions||0;
    ivtKey[k].fraud += r.fraud||0;
  });
  const rows = rev.map(r=>{
    const ivt = ivtKey[r.date + "|" + r.property_tab];
    return {
      date: r.date,
      oem_group: r.oem_group,
      property_label: r.property_label,
      ad_revenue: r.ad_revenue,
      page_views: r.page_views,
      ad_clicks: r.ad_clicks,
      ctr: r.ctr,
      ad_rpm: r.ad_rpm,
      ad_vrpm: r.ad_vrpm,
      impressions: ivt ? ivt.impressions : null,
      fraud_pct: ivt && ivt.impressions ? ivt.fraud/ivt.impressions : null,
    };
  });
  const sorted = sortRows(rows, sortState.detailTable);
  const limit = 500;
  const slice = sorted.slice(0, limit);
  document.getElementById("detailLimit").textContent = limit;
  document.querySelector("#detailTable tbody").innerHTML = slice.length ? slice.map(r=>`
    <tr>
      <td>${r.date}</td>
      <td><span class="pill muted">${r.oem_group}</span></td>
      <td>${r.property_label}</td>
      <td class="num">${fmt.money(r.ad_revenue)}</td>
      <td class="num">${fmt.int(r.page_views)}</td>
      <td class="num">${fmt.int(r.ad_clicks)}</td>
      <td class="num">${fmt.pct(r.ctr)}</td>
      <td class="num">${fmt.num(r.ad_rpm,2)}</td>
      <td class="num">${fmt.num(r.ad_vrpm,2)}</td>
      <td class="num">${fmt.int(r.impressions)}</td>
      <td class="num">${pillFraud(r.fraud_pct)}</td>
    </tr>`).join("") : `<tr><td colspan="11" class="empty">No data in selected filters.</td></tr>`;
}

function pillFraud(p) {
  if (p==null) return "—";
  const cls = p>0.2 ? "bad" : p>0.05 ? "warn" : "good";
  return `<span class="pill ${cls}">${(p*100).toFixed(2)}%</span>`;
}
function flagPill(f) {
  if (f==="bad") return `<span class="pill bad">High fraud</span>`;
  if (f==="warn") return `<span class="pill warn">Watch</span>`;
  if (f==="good") return `<span class="pill good">OK</span>`;
  return `<span class="pill muted">No IVT match</span>`;
}
// PCS = Publisher Conversion Score. Color bands:
//   < 0.3        red    (underperforming)
//   0.3 – < 0.5  orange (borderline)
//   0.5 – 2.1    green  (healthy)
//   > 2.1        purple (suspiciously high — possible inflation / anomaly)
function pcsCell(v) {
  if (v==null) return `<span style="color:var(--muted);">—</span>`;
  const n = Number(v);
  let cls;
  if (n < 0.3) cls = "bad";
  else if (n < 0.5) cls = "orange";
  else if (n <= 2.1) cls = "good";
  else cls = "purple";
  return `<span class="pill ${cls}">${n.toFixed(2)}</span>`;
}

// PCS period badge shown next to the "By property" heading, plus column-header
// labels that reflect the actual date ranges in the current report.
(function renderPcsBadge(){
  const el = document.getElementById("pcsWeekBadge");
  const pcs = DATA.pcs || {};
  const url = pcs.source_url;
  if (el) {
    if (!pcs.period_current && !pcs.period_prior) {
      el.textContent = "PCS: no data";
    } else {
      const parts = [];
      if (pcs.period_current) parts.push(`Current: ${pcs.period_current}`);
      if (pcs.period_prior)   parts.push(`Prior: ${pcs.period_prior}`);
      const label = `PCS — ${parts.join(" · ")}`;
      el.innerHTML = url
        ? `<a href="${url}" target="_blank" style="color:var(--accent);text-decoration:none;">${label}</a>`
        : label;
    }
  }
  // Update the header cells in the property breakout to show the period text.
  const h1 = document.getElementById("pcsCurrentHeader");
  const h2 = document.getElementById("pcsPriorHeader");
  if (h1 && pcs.period_current) h1.textContent = `PCS ${pcs.period_current}`;
  if (h2 && pcs.period_prior)   h2.textContent = `PCS ${pcs.period_prior}`;
})();

// initial render — default to month-to-date
applyPreset("mtd");

// ============================================================
// Tab navigation + Performance tab
// ============================================================
let _perfInit = null;  // becomes a function after first activation
document.querySelectorAll(".tab-btn").forEach(btn => {
  btn.addEventListener("click", () => {
    document.querySelectorAll(".tab-btn").forEach(b => b.classList.toggle("active", b === btn));
    document.querySelectorAll(".tab-content").forEach(c => {
      c.hidden = c.id !== ("tab-" + btn.dataset.tab);
    });
    if (btn.dataset.tab === "performance" && _perfInit) {
      // Charts created while hidden often size wrong; resize after show
      _perfInit();
    }
  });
});

// ---------- Performance tab (IIFE — its own scope) ----------
(function PerfTab() {
  const ALL_PAIRS = DATA.performance_pairs || [];
  const PERF_MIN = (DATA.performance_date_range || {}).min;
  const PERF_MAX = (DATA.performance_date_range || {}).max;
  if (!ALL_PAIRS.length) {
    document.getElementById("perfVerdict").innerHTML =
      "<span style='color:var(--muted);'>No matched IVT × revenue records — Performance tab needs domains in the IVT report to map to property tabs.</span>";
    return;
  }

  const pfmt = {
    pct: v => v==null||Number.isNaN(v)?"—":(v*100).toFixed(2)+"%",
    money: v => v==null||Number.isNaN(v)?"—":"$"+Number(v).toLocaleString(undefined,{maximumFractionDigits:0}),
    int: v => v==null||Number.isNaN(v)?"—":Math.round(v).toLocaleString(),
    r: v => v==null||Number.isNaN(v)?"—":(v>=0?"+":"")+v.toFixed(3),
    p: v => v==null||Number.isNaN(v)?"—":v<0.001?"<0.001":v.toFixed(3),
  };
  function pearson(xs, ys) {
    const n = xs.length;
    if (n < 3) return [NaN, n];
    let sx=0, sy=0;
    for (let i=0;i<n;i++){ sx+=xs[i]; sy+=ys[i]; }
    const mx=sx/n, my=sy/n;
    let num=0, dx=0, dy=0;
    for (let i=0;i<n;i++){
      const a=xs[i]-mx, b=ys[i]-my;
      num+=a*b; dx+=a*a; dy+=b*b;
    }
    if (dx===0 || dy===0) return [NaN, n];
    return [num / Math.sqrt(dx*dy), n];
  }
  function erf(x) {
    const sign = x < 0 ? -1 : 1; x = Math.abs(x);
    const a1=0.254829592,a2=-0.284496736,a3=1.421413741,a4=-1.453152027,a5=1.061405429,p=0.3275911;
    const t=1/(1+p*x);
    return sign*(1-(((((a5*t+a4)*t)+a3)*t+a2)*t+a1)*t*Math.exp(-x*x));
  }
  function tPvalue(r, n) {
    if (n < 4 || Number.isNaN(r) || Math.abs(r) >= 1) return NaN;
    const df = n - 2;
    const t = r * Math.sqrt(df / (1 - r*r));
    const z = t * (1 - 1/(4*df)) / Math.sqrt(1 + t*t/(2*df));
    return 2 * (1 - 0.5 * (1 + erf(Math.abs(z) / Math.SQRT2)));
  }
  function classify(r) {
    if (r==null || Number.isNaN(r)) return {pill:"neu", label:"N/A"};
    const a = Math.abs(r);
    if (a < 0.1) return {pill:"neu", label:"essentially none"};
    if (a < 0.3) return {pill:r<0?"neg":"pos", label:(r<0?"weak negative":"weak positive")};
    if (a < 0.5) return {pill:r<0?"neg":"pos", label:(r<0?"moderate negative":"moderate positive")};
    return {pill:r<0?"neg":"pos", label:(r<0?"strong negative":"strong positive")};
  }
  const BIN_SPECS = [
    [0.00, 0.01,  "0–1%"],
    [0.01, 0.025, "1–2.5%"],
    [0.025, 0.05, "2.5–5%"],
    [0.05, 0.10,  "5–10%"],
    [0.10, 0.20,  "10–20%"],
    [0.20, 1.01,  "20%+"],
  ];
  function binFor(fp) {
    for (const [lo, hi, label] of BIN_SPECS) if (fp >= lo && fp < hi) return label;
    return null;
  }

  function computeAnalysis(joined) {
    const dates = [...new Set(joined.map(r => r.date))].sort();
    const byDay = {};
    for (const r of joined) {
      const v = byDay[r.date] || (byDay[r.date] = {ad_revenue:0, page_views:0, impressions:0, fraud:0});
      v.ad_revenue += r.ad_revenue || 0; v.page_views += r.page_views || 0;
      v.impressions += r.impressions || 0; v.fraud += r.fraud || 0;
    }
    const network_daily = [];
    for (const d of Object.keys(byDay).sort()) {
      const v = byDay[d];
      if (v.page_views < 1000 || v.impressions < 100) continue;
      network_daily.push({date:d, rpm:(v.ad_revenue/v.page_views)*1000,
        fraud_pct:v.fraud/v.impressions, page_views:v.page_views, impressions:v.impressions});
    }
    const [nd_r, nd_n] = pearson(network_daily.map(x=>x.fraud_pct), network_daily.map(x=>x.rpm));

    const byProp = {};
    for (const r of joined) {
      const p = byProp[r.property_tab] || (byProp[r.property_tab] = {
        tab:r.property_tab, label:r.property_label, oem_group:r.oem_group,
        ad_revenue:0, page_views:0, ad_clicks:0, impressions:0, fraud:0, days:0});
      p.ad_revenue += r.ad_revenue || 0; p.page_views += r.page_views || 0;
      p.ad_clicks += r.ad_clicks || 0; p.impressions += r.impressions || 0;
      p.fraud += r.fraud || 0; p.days += 1;
    }
    const per_property = [];
    const pcsByProp = (DATA.pcs && DATA.pcs.by_property) || {};
    for (const tab of Object.keys(byProp)) {
      const v = byProp[tab];
      const minDays = Math.min(5, dates.length);
      if (v.days < minDays || v.page_views < 1000 || v.impressions < 100) continue;
      const pcs = pcsByProp[v.tab] || {};
      per_property.push({
        tab:v.tab, label:v.label, oem_group:v.oem_group, days:v.days,
        rpm:(v.ad_revenue/v.page_views)*1000,
        ctr: v.page_views ? v.ad_clicks/v.page_views : null,
        fraud_pct: v.fraud/v.impressions,
        ad_revenue:v.ad_revenue, page_views:v.page_views,
        ad_clicks:v.ad_clicks, impressions:v.impressions,
        pcs_current: (pcs.pcs_current == null ? null : pcs.pcs_current),
        pcs_prior:   (pcs.pcs_prior   == null ? null : pcs.pcs_prior),
      });
    }
    for (const p of per_property) {
      p.cpc = (p.ad_clicks > 0) ? p.ad_revenue / p.ad_clicks : null;
      p.cpc_clean = p.cpc != null ? p.cpc * (1 - p.fraud_pct) : null;
    }
    if (per_property.length >= 3) {
      const sorted = per_property.map(p=>p.cpc_clean).filter(v=>v!=null && !Number.isNaN(v)).sort((a,b)=>a-b);
      const t1 = sorted[Math.floor(sorted.length*1/3)];
      const t2 = sorted[Math.floor(sorted.length*2/3)];
      for (const p of per_property) {
        if (p.cpc_clean == null) { p.demand_tier = "—"; p.demand_why = "No clicks recorded."; }
        else if (p.cpc_clean >= t2) {
          p.demand_tier = "High";
          p.demand_why = `CPC $${p.cpc.toFixed(3)} × (1 − fraud ${(p.fraud_pct*100).toFixed(1)}%) = $${p.cpc_clean.toFixed(3)} clean CPC — top tertile.`;
        } else if (p.cpc_clean >= t1) {
          p.demand_tier = "Mid";
          p.demand_why = `Clean CPC $${p.cpc_clean.toFixed(3)} — middle tertile.`;
        } else {
          p.demand_tier = "Low";
          p.demand_why = `Clean CPC $${p.cpc_clean.toFixed(3)} — bottom tertile.`;
        }
      }
    } else for (const p of per_property) { p.demand_tier = "—"; p.demand_why = ""; }
    for (const p of per_property) {
      p.focus_score = (p.cpc_clean != null) ? p.cpc_clean * Math.log10(Math.max(10, p.page_views || 0)) : null;
    }
    const [pa_r, pa_n] = pearson(per_property.map(x=>x.fraud_pct), per_property.map(x=>x.rpm));

    const propRecs = {};
    for (const r of joined) (propRecs[r.property_tab] || (propRecs[r.property_tab] = [])).push(r);
    const within_xs = [], within_ys = [];
    const per_property_corr = [];
    for (const tab of Object.keys(propRecs)) {
      const recs = propRecs[tab];
      const minN = Math.max(3, Math.min(7, dates.length));
      if (recs.length < minN) continue;
      const fp = recs.map(r=>r.fraud_pct), rpm = recs.map(r=>r.rpm);
      const mfp = fp.reduce((a,b)=>a+b,0)/fp.length, mrpm = rpm.reduce((a,b)=>a+b,0)/rpm.length;
      for (let i=0;i<fp.length;i++){ within_xs.push(fp[i]-mfp); within_ys.push(rpm[i]-mrpm); }
      const [r_p, n_p] = pearson(fp, rpm);
      per_property_corr.push({tab, label:recs[0].property_label, oem_group:recs[0].oem_group, n:n_p, r:r_p, mean_rpm:mrpm, mean_fraud:mfp});
    }
    const [wi_r, wi_n] = pearson(within_xs, within_ys);
    per_property_corr.sort((a,b)=>{
      const ax = Number.isNaN(a.r) ? -1 : Math.abs(a.r);
      const bx = Number.isNaN(b.r) ? -1 : Math.abs(b.r);
      return bx - ax;
    });

    const cross_bins = BIN_SPECS.map(([_,__,label]) => ({bin:label, n:0, rpms:[], page_views:0, ad_revenue:0, impressions:0, fraud:0}));
    const bIdx = Object.fromEntries(cross_bins.map((b,i)=>[b.bin,i]));
    for (const r of joined) {
      const b = binFor(r.fraud_pct); if (!b) continue;
      const x = cross_bins[bIdx[b]];
      x.rpms.push(r.rpm); x.n += 1;
      x.page_views += r.page_views || 0; x.ad_revenue += r.ad_revenue || 0;
      x.impressions += r.impressions || 0; x.fraud += r.fraud || 0;
    }
    for (const b of cross_bins) {
      if (!b.n) { b.weighted_rpm = null; b.mean_rpm = null; b.median_rpm = null; continue; }
      const sorted = b.rpms.slice().sort((a,b)=>a-b);
      b.mean_rpm = sorted.reduce((a,b)=>a+b,0)/b.n;
      b.median_rpm = b.n % 2 ? sorted[(b.n-1)/2] : (sorted[b.n/2-1]+sorted[b.n/2])/2;
      b.weighted_rpm = b.page_views ? (b.ad_revenue/b.page_views)*1000 : null;
      delete b.rpms;
    }
    const baseline_w = (cross_bins.find(b => b.n && b.weighted_rpm) || {}).weighted_rpm || null;
    for (const b of cross_bins) {
      if (b.n && baseline_w && b.weighted_rpm != null) b.pct_of_baseline = b.weighted_rpm / baseline_w;
    }
    const propMeanRpm = {};
    for (const tab of Object.keys(propRecs)) {
      const recs = propRecs[tab]; if (recs.length < 5) continue;
      propMeanRpm[tab] = recs.reduce((a,r)=>a+r.rpm,0) / recs.length;
    }
    const within_bins = BIN_SPECS.map(([_,__,label]) => ({bin:label, n:0, vals:[]}));
    const wbIdx = Object.fromEntries(within_bins.map((b,i)=>[b.bin,i]));
    for (const r of joined) {
      const m = propMeanRpm[r.property_tab]; if (!m || m <= 0) continue;
      const b = binFor(r.fraud_pct); if (!b) continue;
      within_bins[wbIdx[b]].vals.push(r.rpm / m);
      within_bins[wbIdx[b]].n += 1;
    }
    for (const b of within_bins) {
      if (!b.n) { b.mean_rel_rpm = null; b.median_rel_rpm = null; continue; }
      const v = b.vals.slice().sort((a,b)=>a-b);
      b.mean_rel_rpm = v.reduce((a,b)=>a+b,0)/b.n;
      b.median_rel_rpm = v[Math.floor(b.n/2)];
      delete b.vals;
    }
    let threshold_bin = null;
    for (const b of cross_bins) {
      if (b.n && b.pct_of_baseline != null && b.pct_of_baseline <= 0.75) { threshold_bin = b.bin; break; }
    }
    return {
      summary:{
        joined_rows:joined.length, properties_used:per_property.length,
        date_range:[dates[0]||null, dates[dates.length-1]||null],
        network_daily:{r:nd_r, n:nd_n, p:tPvalue(nd_r, nd_n)},
        per_property_aggregate:{r:pa_r, n:pa_n, p:tPvalue(pa_r, pa_n)},
        within_property_pooled:{r:wi_r, n:wi_n, p:tPvalue(wi_r, wi_n)},
      },
      network_daily, per_property, per_property_corr,
      threshold:{bins:cross_bins, within_bins, baseline_weighted_rpm:baseline_w, threshold_bin},
    };
  }

  // ---- sortable helper ----
  function makeSorter(tableId, defaultKey, defaultDir, renderFn) {
    const state = {k:defaultKey, dir:defaultDir};
    function paintHeader() {
      document.querySelectorAll(`#${tableId} th[data-k]`).forEach(th=>{
        const arr = th.querySelector(".arr") || (() => {
          const s = document.createElement("span"); s.className = "arr"; s.textContent = "↕";
          s.style.marginLeft = "4px"; s.style.fontSize = "10px"; s.style.opacity = "0.4";
          th.appendChild(s); return s;
        })();
        th.classList.remove("sort-asc","sort-desc");
        if (th.dataset.k === state.k) {
          th.classList.add(state.dir > 0 ? "sort-asc" : "sort-desc");
          arr.textContent = state.dir > 0 ? "▲" : "▼"; arr.style.opacity = "1";
        } else { arr.textContent = "↕"; arr.style.opacity = "0.4"; }
      });
    }
    document.querySelectorAll(`#${tableId} th[data-k]`).forEach(th=>{
      th.style.cursor = "pointer";
      th.addEventListener("click", ()=>{
        const k = th.dataset.k;
        if (state.k === k) state.dir = -state.dir;
        else { state.k = k; state.dir = -1; }
        paintHeader(); renderFn(state);
      });
    });
    paintHeader();
    return () => renderFn(state);
  }
  function sortRows(rows, key, dir) {
    return rows.slice().sort((a,b)=>{
      const x = a[key], y = b[key];
      if (x == null && y == null) return 0;
      if (x == null) return 1;
      if (y == null) return -1;
      if (typeof x === "number" && typeof y === "number") return (x - y) * dir;
      return String(x).localeCompare(String(y)) * dir;
    });
  }

  function tierPill(tier, why) {
    const t = why ? ` title="${why.replace(/"/g,"&quot;")}"` : "";
    if (tier === "High") return `<span class="pill pos"${t}>High</span>`;
    if (tier === "Mid")  return `<span class="pill" style="background:rgba(251,191,36,0.18);color:var(--warn)"${t}>Mid</span>`;
    if (tier === "Low")  return `<span class="pill" style="background:rgba(139,147,167,0.22);color:var(--muted)"${t}>Low</span>`;
    return `<span class="pill muted">—</span>`;
  }

  // ---- charts ----
  const pcharts = {};
  function buildCharts() {
    const baselinePlugin = {
      id:"perfBaseline",
      afterDraw:(chart)=>{
        const baseline = chart.options.plugins?.baselineValue;
        if (baseline == null) return;
        const {ctx, chartArea, scales} = chart;
        const y = scales.y.getPixelForValue(baseline);
        ctx.save(); ctx.strokeStyle = "rgba(255,255,255,0.4)"; ctx.setLineDash([4,4]);
        ctx.beginPath(); ctx.moveTo(chartArea.left, y); ctx.lineTo(chartArea.right, y); ctx.stroke();
        ctx.fillStyle = "rgba(255,255,255,0.6)"; ctx.font = "11px sans-serif";
        ctx.fillText(`baseline $${baseline.toFixed(2)}`, chartArea.left+4, y-4);
        ctx.restore();
      }
    };
    const unitLinePlugin = {
      id:"perfUnitLine",
      afterDraw:(chart)=>{
        if (!chart.options.plugins?.unitLine) return;
        const {ctx, chartArea, scales} = chart;
        const y = scales.y.getPixelForValue(1);
        ctx.save(); ctx.strokeStyle = "rgba(255,255,255,0.4)"; ctx.setLineDash([4,4]);
        ctx.beginPath(); ctx.moveTo(chartArea.left, y); ctx.lineTo(chartArea.right, y); ctx.stroke();
        ctx.fillStyle = "rgba(255,255,255,0.6)"; ctx.font = "11px sans-serif";
        ctx.fillText("property's typical RPM", chartArea.left+4, y-4);
        ctx.restore();
      }
    };
    pcharts.bin = new Chart(document.getElementById("perfBinChart"), {
      type:"bar",
      data:{labels:[], datasets:[{data:[], backgroundColor:[], borderColor:"rgba(255,255,255,0.15)", borderWidth:1}]},
      options:{plugins:{legend:{display:false}, tooltip:{callbacks:{label:(c)=>c.chart.options.plugins.binTooltips?.[c.dataIndex] || ""}}},
        scales:{y:{title:{display:true,text:"Weighted RPM ($)"}, ticks:{callback:v=>"$"+v.toFixed(2)}, grid:{color:"rgba(255,255,255,0.05)"}},
          x:{title:{display:true,text:"Fraud % bucket"}}},
        maintainAspectRatio:false, responsive:true},
      plugins: [baselinePlugin],
    });
    pcharts.withinBin = new Chart(document.getElementById("perfWithinBinChart"), {
      type:"bar",
      data:{labels:[], datasets:[{data:[], backgroundColor:[], borderColor:"rgba(255,255,255,0.15)", borderWidth:1}]},
      options:{plugins:{legend:{display:false}, unitLine:true,
          tooltip:{callbacks:{label:(c)=>c.chart.options.plugins.withinTooltips?.[c.dataIndex] || ""}}},
        scales:{y:{title:{display:true,text:"RPM relative to property's avg"}, ticks:{callback:v=>v.toFixed(2)+"×"},
            grid:{color:"rgba(255,255,255,0.05)"}, suggestedMin:0, suggestedMax:1.5},
          x:{title:{display:true,text:"Fraud % bucket"}}},
        maintainAspectRatio:false, responsive:true},
      plugins: [unitLinePlugin],
    });
    pcharts.ndScatter = new Chart(document.getElementById("perfNdScatter"), {
      type:"scatter",
      data:{datasets:[{label:"Network day", data:[], backgroundColor:"rgba(106,169,255,0.55)", borderColor:"rgba(106,169,255,0.9)", pointRadius:4}]},
      options:{plugins:{legend:{display:false}},
        scales:{x:{title:{display:true,text:"Network Fraud %"}, ticks:{callback:v=>(v*100).toFixed(0)+"%"}},
          y:{title:{display:true,text:"Network RPM"}, ticks:{callback:v=>"$"+v.toFixed(2)}}},
        maintainAspectRatio:false, responsive:true}});
    pcharts.ndTime = new Chart(document.getElementById("perfNdTime"), {
      type:"line",
      data:{labels:[], datasets:[
        {label:"Fraud %", data:[], borderColor:"#f87171", backgroundColor:"rgba(248,113,113,0.12)", yAxisID:"y", tension:0.25, fill:true},
        {label:"RPM", data:[], borderColor:"#6aa9ff", backgroundColor:"transparent", yAxisID:"y1", tension:0.25, fill:false}]},
      options:{plugins:{legend:{labels:{boxWidth:12}}},
        scales:{y:{position:"left", ticks:{callback:v=>(v*100).toFixed(0)+"%"}, title:{display:true,text:"Fraud %"}},
          y1:{position:"right", grid:{drawOnChartArea:false}, ticks:{callback:v=>"$"+v.toFixed(2)}, title:{display:true,text:"RPM"}}},
        maintainAspectRatio:false, responsive:true}});
    pcharts.demandScatter = new Chart(document.getElementById("perfDemandScatter"), {
      type:"bubble",
      data:{datasets:[{label:"Property", data:[], backgroundColor:[], borderColor:[], borderWidth:1.5}]},
      options:{plugins:{legend:{display:false},
          tooltip:{callbacks:{label:c=>{
            const r = c.raw;
            return [`${r.label}`, `Clean CPC: ${r.x!=null?'$'+r.x.toFixed(3):'—'}  (raw $${r.cpc.toFixed(3)})`,
              `Volume: ${Math.round(r.y).toLocaleString()} page views`, `Revenue: $${Math.round(r.rev).toLocaleString()}`,
              `Fraud: ${(r.fraud*100).toFixed(2)}%`, `CTR: ${r.ctr!=null?(r.ctr*100).toFixed(2)+'%':'—'}`,
              `PCS current: ${r.pcs_current!=null?r.pcs_current.toFixed(2):'—'}  ·  prior: ${r.pcs_prior!=null?r.pcs_prior.toFixed(2):'—'}`,
              `Demand tier: ${r.tier}`];
          }}}},
        scales:{x:{title:{display:true,text:"Clean CPC ($/click)"}, ticks:{callback:v=>"$"+v.toFixed(2)}},
          y:{title:{display:true,text:"Page views (volume)"}, ticks:{callback:v=>v>=1000?(v/1000).toFixed(0)+"k":v.toLocaleString()}}},
        maintainAspectRatio:false, responsive:true}});
    pcharts.ppScatter = new Chart(document.getElementById("perfPpScatter"), {
      type:"bubble",
      data:{datasets:[{label:"Property", data:[], backgroundColor:[], borderColor:[], borderWidth:1.5}]},
      options:{plugins:{legend:{display:false},
          tooltip:{callbacks:{label:c=>{
            const r = c.raw;
            return [`${r.label}`, `RPM: $${r.y.toFixed(2)}`, `Fraud: ${(r.x*100).toFixed(2)}%`,
              `CTR: ${r.ctr!=null?(r.ctr*100).toFixed(2)+'%':'—'}`,
              `Volume: ${Math.round(r.pv).toLocaleString()} page views`, `Tier: ${r.tier}`];
          }}}},
        scales:{x:{title:{display:true,text:"Fraud %"}, ticks:{callback:v=>(v*100).toFixed(0)+"%"}},
          y:{title:{display:true,text:"RPM"}, ticks:{callback:v=>"$"+v.toFixed(2)}}},
        maintainAspectRatio:false, responsive:true}});
  }

  // ---- render functions ----
  function renderVerdict(c) {
    const pa = c.summary.per_property_aggregate, wi = c.summary.within_property_pooled;
    const paC = classify(pa.r);
    let lead;
    if (Number.isNaN(pa.r)) lead = `<span style="font-weight:700;color:var(--muted);">Not enough data in this date range to compute a meaningful correlation.</span>`;
    else if (Math.abs(pa.r) < 0.1) lead = `<span style="font-weight:700;color:var(--muted);">In this window, IVT fraud % and RPM are essentially uncorrelated.</span>`;
    else if (pa.r < 0) lead = `<span style="font-weight:700;color:var(--bad);">Negative correlation: properties with higher fraud have lower RPM.</span>`;
    else lead = `<span style="font-weight:700;color:var(--good);">Positive correlation: properties with higher fraud have higher RPM (unusual — investigate).</span>`;
    const wiText = Number.isNaN(wi.r) ? "n/a (need ≥4 days)" : `r = ${pfmt.r(wi.r)}`;
    const color = paC.pill==='neg'?'var(--bad)':paC.pill==='pos'?'var(--good)':'var(--muted)';
    document.getElementById("perfVerdict").innerHTML = `${lead}<br><br>
      <span style="font-size:14px;">
        Across properties (n = ${pa.n}), <span style="color:${color};font-weight:600;">${paC.label}, r = ${pfmt.r(pa.r)}, p ${pa.p<0.001?"<":"="} ${pfmt.p(pa.p)}</span>.
        Within properties (each centered to its own mean), pooled ${wiText}.
      </span>`;
  }
  function renderLevels(c) {
    const lv = c.summary;
    function lc(label, x, hint) {
      const cls = classify(x.r);
      const color = cls.pill==='neg'?'var(--bad)':cls.pill==='pos'?'var(--good)':'var(--text)';
      return `<div style="background:var(--panel2);border:1px solid var(--line);border-radius:10px;padding:14px;">
        <div style="font-size:11px;text-transform:uppercase;color:var(--muted);letter-spacing:0.05em;">${label}</div>
        <div style="font-size:28px;font-weight:700;margin-top:6px;font-variant-numeric:tabular-nums;color:${color};">r = ${pfmt.r(x.r)}</div>
        <div style="font-size:12px;color:var(--muted);margin-top:4px;">n = ${x.n.toLocaleString()} · p ${x.p<0.001?"<":"="} ${pfmt.p(x.p)}<br>${hint}</div>
      </div>`;
    }
    document.getElementById("perfLevels").innerHTML = [
      lc("Network daily (pooled days)", lv.network_daily, "Each day is one observation."),
      lc("Per-property aggregate (one row per property)", lv.per_property_aggregate, "Aggregated within window."),
      lc("Within-property pooled (strictest)", lv.within_property_pooled, "Each property centered to its own mean."),
    ].join("");
  }
  function renderThresholdHeadline(c) {
    const t = c.threshold, bins = t.bins, filled = bins.filter(b => b.n);
    let head;
    if (!filled.length) head = `<span style="color:var(--muted);">No data in this date range.</span>`;
    else if (t.threshold_bin) head = `<strong>RPM drops sharply once fraud crosses <span style="color:var(--bad);">${t.threshold_bin}</span>.</strong>
      Below that, RPM stays within ~25% of the baseline; above it, it falls off.`;
    else head = `RPM does not drop more than 25% below the baseline in any of the observed fraud buckets in this window.`;
    if (filled.length >= 2) {
      const first = filled[0], last = filled[filled.length-1];
      const dropPct = first.weighted_rpm ? (1 - (last.weighted_rpm || 0)/first.weighted_rpm) * 100 : null;
      head += `<br><br><span style="color:var(--muted);font-size:14px;">
        Cleanest bucket (<em>${first.bin}</em>): weighted RPM <strong>$${first.weighted_rpm.toFixed(2)}</strong>;
        most fraud-heavy (<em>${last.bin}</em>): <strong>${last.weighted_rpm?'$'+last.weighted_rpm.toFixed(2):'—'}</strong>${dropPct!=null?` — a ${dropPct.toFixed(0)}% drop`:''}.
      </span>`;
    }
    document.getElementById("perfThresholdHeadline").innerHTML = head;
  }
  function renderPerfCharts(c) {
    // Bin
    const bins = c.threshold.bins, baseline = c.threshold.baseline_weighted_rpm;
    pcharts.bin.data.labels = bins.map(b => b.bin);
    pcharts.bin.data.datasets[0].data = bins.map(b => b.weighted_rpm || 0);
    pcharts.bin.data.datasets[0].backgroundColor = bins.map(b => {
      if (!b.weighted_rpm || !baseline) return "rgba(139,147,167,0.5)";
      const ratio = b.weighted_rpm / baseline;
      if (ratio >= 0.9) return "rgba(74,222,128,0.7)";
      if (ratio >= 0.75) return "rgba(251,191,36,0.7)";
      return "rgba(248,113,113,0.7)";
    });
    pcharts.bin.options.plugins.baselineValue = baseline;
    pcharts.bin.options.plugins.binTooltips = bins.map(b => {
      if (!b.n) return "no data";
      const pct = b.pct_of_baseline ? `${(b.pct_of_baseline*100).toFixed(0)}% of baseline` : "";
      return `Weighted RPM: $${b.weighted_rpm.toFixed(2)} · n: ${b.n} · ${pct}`;
    });
    pcharts.bin.update();

    const wb = c.threshold.within_bins;
    pcharts.withinBin.data.labels = wb.map(b => b.bin);
    pcharts.withinBin.data.datasets[0].data = wb.map(b => b.mean_rel_rpm || 0);
    pcharts.withinBin.data.datasets[0].backgroundColor = wb.map(b => {
      if (!b.mean_rel_rpm) return "rgba(139,147,167,0.5)";
      if (b.mean_rel_rpm >= 0.95) return "rgba(74,222,128,0.7)";
      if (b.mean_rel_rpm >= 0.8) return "rgba(251,191,36,0.7)";
      return "rgba(248,113,113,0.7)";
    });
    pcharts.withinBin.options.plugins.withinTooltips = wb.map(b => {
      if (!b.n) return "no data";
      return `Mean: ${b.mean_rel_rpm.toFixed(2)}× (${(b.mean_rel_rpm*100).toFixed(0)}% of property avg) · n: ${b.n}`;
    });
    pcharts.withinBin.update();

    pcharts.ndScatter.data.datasets[0].data = c.network_daily.map(r => ({x:r.fraud_pct, y:r.rpm}));
    pcharts.ndScatter.update();

    const sortedDays = c.network_daily.slice().sort((a,b)=>a.date.localeCompare(b.date));
    pcharts.ndTime.data.labels = sortedDays.map(r => r.date);
    pcharts.ndTime.data.datasets[0].data = sortedDays.map(r => r.fraud_pct);
    pcharts.ndTime.data.datasets[1].data = sortedDays.map(r => r.rpm);
    pcharts.ndTime.update();

    // Demand bubble
    {
      const props = c.per_property;
      const revenues = props.map(p=>p.ad_revenue).filter(v=>v>0);
      const maxRev = revenues.length ? Math.max(...revenues) : 1;
      const minR=5, maxR=28;
      function fc(fp) {
        if (fp == null || Number.isNaN(fp)) return ["rgba(139,147,167,0.55)","rgba(139,147,167,0.95)"];
        if (fp >= 0.20) return ["rgba(248,113,113,0.55)","rgba(248,113,113,0.95)"];
        if (fp >= 0.05) return ["rgba(251,191,36,0.55)","rgba(251,191,36,0.95)"];
        return ["rgba(74,222,128,0.55)","rgba(74,222,128,0.95)"];
      }
      const data = props.filter(p=>p.cpc_clean!=null).map(p=>{
        const r = p.ad_revenue > 0 ? minR + (maxR-minR)*Math.sqrt(p.ad_revenue/maxRev) : minR;
        return {x:p.cpc_clean, y:p.page_views, r, label:p.label, cpc:p.cpc, rev:p.ad_revenue,
          fraud:p.fraud_pct, ctr:p.ctr, tier:p.demand_tier, pcs_current:p.pcs_current, pcs_prior:p.pcs_prior};
      });
      pcharts.demandScatter.data.datasets[0].data = data;
      pcharts.demandScatter.data.datasets[0].backgroundColor = data.map(d=>fc(d.fraud)[0]);
      pcharts.demandScatter.data.datasets[0].borderColor = data.map(d=>fc(d.fraud)[1]);
      pcharts.demandScatter.update();
    }

    // PP fraud×RPM bubble
    {
      const props = c.per_property;
      const ctrs = props.map(p=>p.ctr).filter(v=>v!=null && !Number.isNaN(v)).sort((a,b)=>a-b);
      const ctrLow = ctrs.length ? ctrs[Math.floor(ctrs.length/3)] : 0;
      const ctrHigh = ctrs.length ? ctrs[Math.floor(ctrs.length*2/3)] : 0;
      const pvs = props.map(p=>p.page_views).filter(v=>v>0);
      const maxPv = pvs.length ? Math.max(...pvs) : 1;
      const minR=5, maxR=26;
      function ctrColor(ctr) {
        if (ctr==null || Number.isNaN(ctr)) return ["rgba(139,147,167,0.55)","rgba(139,147,167,0.95)"];
        if (ctr <= ctrLow) return ["rgba(248,113,113,0.55)","rgba(248,113,113,0.95)"];
        if (ctr >= ctrHigh) return ["rgba(74,222,128,0.55)","rgba(74,222,128,0.95)"];
        return ["rgba(251,191,36,0.55)","rgba(251,191,36,0.95)"];
      }
      const data = props.map(p=>{
        const r = p.page_views > 0 ? minR + (maxR-minR)*Math.sqrt(p.page_views/maxPv) : minR;
        return {x:p.fraud_pct, y:p.rpm, r, label:p.label, ctr:p.ctr, pv:p.page_views, tier:p.demand_tier};
      });
      pcharts.ppScatter.data.datasets[0].data = data;
      pcharts.ppScatter.data.datasets[0].backgroundColor = props.map(p=>ctrColor(p.ctr)[0]);
      pcharts.ppScatter.data.datasets[0].borderColor = props.map(p=>ctrColor(p.ctr)[1]);
      pcharts.ppScatter.update();
    }
  }

  // ---- sortable tables ----
  let pcomputed = null;
  const reBin = makeSorter("perfBinTable", "bin_order", 1, (state)=>{
    if (!pcomputed) return;
    const bins = pcomputed.threshold.bins.map((b,i)=>({...b, bin_order:i}));
    const sorted = sortRows(bins, state.k, state.dir);
    document.querySelector("#perfBinTable tbody").innerHTML = sorted.map(b=>{
      if (!b.n) return `<tr><td>${b.bin}</td><td class="num" colspan="7" style="color:var(--muted);font-style:italic;">no data</td></tr>`;
      const pct = b.pct_of_baseline;
      const pill = pct==null ? "" :
        pct>=0.9 ? `<span class="pill pos">${(pct*100).toFixed(0)}%</span>` :
        pct>=0.75 ? `<span class="pill" style="background:rgba(251,191,36,0.18);color:var(--warn)">${(pct*100).toFixed(0)}%</span>` :
        `<span class="pill neg">${(pct*100).toFixed(0)}%</span>`;
      return `<tr><td><strong>${b.bin}</strong></td>
        <td class="num">${b.n.toLocaleString()}</td>
        <td class="num">${b.weighted_rpm?"$"+b.weighted_rpm.toFixed(2):"—"}</td>
        <td class="num">${b.mean_rpm?"$"+b.mean_rpm.toFixed(2):"—"}</td>
        <td class="num">${b.median_rpm?"$"+b.median_rpm.toFixed(2):"—"}</td>
        <td class="num">${pill}</td>
        <td class="num">${Math.round(b.page_views).toLocaleString()}</td>
        <td class="num">$${Math.round(b.ad_revenue).toLocaleString()}</td></tr>`;
    }).join("");
  });
  const reProp = makeSorter("perfPropTable", "focus_score", -1, (state)=>{
    if (!pcomputed) return;
    const rows = sortRows(pcomputed.per_property, state.k, state.dir);
    document.querySelector("#perfPropTable tbody").innerHTML = rows.length ? rows.map(r=>`
      <tr><td><span class="pill muted">${r.oem_group}</span></td>
        <td>${r.label}</td>
        <td class="num">${r.days}</td>
        <td class="num">${r.cpc==null||Number.isNaN(r.cpc)?"—":"$"+r.cpc.toFixed(3)}</td>
        <td class="num">${r.cpc_clean==null||Number.isNaN(r.cpc_clean)?"—":"$"+r.cpc_clean.toFixed(3)}</td>
        <td class="num">${r.rpm==null||Number.isNaN(r.rpm)?"—":"$"+r.rpm.toFixed(2)}</td>
        <td class="num">${pfmt.pct(r.ctr)}</td>
        <td class="num">${pfmt.pct(r.fraud_pct)}</td>
        <td class="num">${pfmt.int(r.page_views)}</td>
        <td class="num">${pfmt.money(r.ad_revenue)}</td>
        <td class="num">${r.focus_score==null?"—":r.focus_score.toFixed(2)}</td>
        <td>${tierPill(r.demand_tier, r.demand_why)}</td></tr>`).join("") :
      `<tr><td colspan="12" style="color:var(--muted);font-style:italic;text-align:center;padding:18px;">No properties met the volume threshold in this window.</td></tr>`;
  });
  const rePpCorr = makeSorter("perfPpCorrTable", "r", 1, (state)=>{
    if (!pcomputed) return;
    const rows = sortRows(pcomputed.per_property_corr, state.k, state.dir);
    document.querySelector("#perfPpCorrTable tbody").innerHTML = rows.length ? rows.map(r=>{
      const cl = classify(r.r);
      return `<tr><td><span class="pill muted">${r.oem_group}</span></td>
        <td>${r.label}</td>
        <td class="num">${r.n}</td>
        <td class="num"><span class="pill ${cl.pill}">${pfmt.r(r.r)}</span></td>
        <td class="num">$${r.mean_rpm.toFixed(2)}</td>
        <td class="num">${pfmt.pct(r.mean_fraud)}</td></tr>`;
    }).join("") : `<tr><td colspan="6" style="color:var(--muted);font-style:italic;text-align:center;padding:18px;">No properties have enough days in this window.</td></tr>`;
  });

  // ---- filter ----
  const pstate = {from:null, to:null};
  function shiftDate(iso, days) {
    const d = new Date(iso + "T00:00:00Z"); d.setUTCDate(d.getUTCDate() + days);
    return d.toISOString().slice(0,10);
  }
  function applyPerfPreset(name) {
    let from, to;
    if (name === "yesterday") { from = to = PERF_MAX; }
    else if (name === "7")    { from = shiftDate(PERF_MAX, -6);  to = PERF_MAX; }
    else if (name === "14")   { from = shiftDate(PERF_MAX, -13); to = PERF_MAX; }
    else if (name === "30")   { from = shiftDate(PERF_MAX, -29); to = PERF_MAX; }
    else if (name === "90")   { from = shiftDate(PERF_MAX, -89); to = PERF_MAX; }
    else                       { from = PERF_MIN; to = PERF_MAX; name = "all"; }
    if (from < PERF_MIN) from = PERF_MIN;
    pstate.from = from; pstate.to = to;
    document.getElementById("perfFrom").value = from;
    document.getElementById("perfTo").value = to;
    document.querySelectorAll("#perfPresets button").forEach(b => b.classList.toggle("active", b.dataset.d === name));
    rerender();
  }
  function rerender() {
    const filtered = ALL_PAIRS.filter(r => (!pstate.from || r.date >= pstate.from) && (!pstate.to || r.date <= pstate.to));
    pcomputed = computeAnalysis(filtered);
    document.getElementById("perfFilterMeta").textContent =
      `${pstate.from} → ${pstate.to}  ·  ${filtered.length.toLocaleString()} property-day records`;
    renderVerdict(pcomputed); renderLevels(pcomputed); renderThresholdHeadline(pcomputed);
    renderPerfCharts(pcomputed);
    renderScaleGrid(pcomputed);
    reBin(); reProp(); rePpCorr();
  }

  // ---- Scale candidates grid: top-half clean CPC + below-median volume + fraud < 10% ----
  function renderScaleGrid(c) {
    const grid = document.getElementById("perfScaleGrid");
    if (!grid) return;
    const props = c.per_property.filter(p => p.cpc_clean != null);
    if (props.length < 4) {
      grid.innerHTML = `<div style="color:var(--muted);font-style:italic;padding:14px;">Not enough properties in this window to identify scale candidates.</div>`;
      return;
    }
    const cpcs = props.map(p => p.cpc_clean).slice().sort((a,b)=>a-b);
    const vols = props.map(p => p.page_views).slice().sort((a,b)=>a-b);
    const medCpc = cpcs[Math.floor(cpcs.length/2)];
    const medVol = vols[Math.floor(vols.length/2)];
    const candidates = props
      .filter(p => p.cpc_clean >= medCpc && p.page_views <= medVol && p.fraud_pct < 0.10)
      .sort((a,b) => b.cpc_clean - a.cpc_clean)
      .slice(0, 8);
    if (!candidates.length) {
      grid.innerHTML = `<div style="color:var(--muted);font-style:italic;padding:14px;">
        No properties match all three criteria (top-half CPC + below-median volume + fraud &lt; 10%) in this window.
        Try widening the date range.
      </div>`;
      return;
    }
    grid.innerHTML = candidates.map(p => {
      const headroomViews = Math.max(0, medVol - p.page_views);
      const impliedAddRev = (p.ctr || 0) * p.cpc_clean * headroomViews;
      const fraudColor = p.fraud_pct < 0.05 ? "var(--good)" :
        p.fraud_pct < 0.10 ? "var(--warn)" : "var(--bad)";
      // PCS current — color matches the property breakout bands.
      let pcsColor = "var(--muted)", pcsText = "—";
      if (p.pcs_current != null) {
        const n = Number(p.pcs_current);
        pcsText = n.toFixed(2);
        if (n < 0.3) pcsColor = "var(--bad)";
        else if (n < 0.5) pcsColor = "#fb923c";      // orange
        else if (n <= 2.1) pcsColor = "var(--good)";
        else pcsColor = "#a78bfa";                    // purple
      }
      const pcsPeriod = (DATA.pcs && DATA.pcs.period_current) ? DATA.pcs.period_current : "current";
      return `<div style="background:var(--panel2);border:1px solid var(--line);border-left:3px solid var(--good);border-radius:8px;padding:14px;">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px;margin-bottom:10px;">
          <div>
            <div style="font-size:13px;font-weight:600;color:var(--text);line-height:1.3;">${p.label}</div>
            <div style="font-size:11px;color:var(--muted);margin-top:2px;">${p.oem_group}</div>
          </div>
          <span class="pill pos" style="font-size:10px;">SCALE</span>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px 12px;font-size:12px;">
          <div>
            <div style="color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:0.05em;">Clean CPC</div>
            <div style="font-weight:600;font-variant-numeric:tabular-nums;">$${p.cpc_clean.toFixed(3)}</div>
          </div>
          <div>
            <div style="color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:0.05em;">CTR</div>
            <div style="font-weight:600;font-variant-numeric:tabular-nums;">${p.ctr!=null?(p.ctr*100).toFixed(2)+"%":"—"}</div>
          </div>
          <div>
            <div style="color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:0.05em;">Volume</div>
            <div style="font-weight:600;font-variant-numeric:tabular-nums;">${Math.round(p.page_views).toLocaleString()}</div>
          </div>
          <div>
            <div style="color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:0.05em;">Fraud</div>
            <div style="font-weight:600;font-variant-numeric:tabular-nums;color:${fraudColor};">${(p.fraud_pct*100).toFixed(2)}%</div>
          </div>
          <div style="grid-column:1 / -1;">
            <div style="color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:0.05em;">PCS · ${pcsPeriod}</div>
            <div style="font-weight:600;font-variant-numeric:tabular-nums;color:${pcsColor};">${pcsText}</div>
          </div>
          <div style="grid-column:1 / -1;border-top:1px solid var(--line);padding-top:8px;margin-top:4px;">
            <div style="color:var(--muted);font-size:10px;text-transform:uppercase;letter-spacing:0.05em;">Current revenue / scaled potential</div>
            <div style="display:flex;justify-content:space-between;align-items:baseline;margin-top:2px;">
              <span style="font-weight:600;font-variant-numeric:tabular-nums;">$${Math.round(p.ad_revenue).toLocaleString()}</span>
              <span style="color:var(--good);font-size:11px;font-variant-numeric:tabular-nums;">+~$${Math.round(impliedAddRev).toLocaleString()} at median volume</span>
            </div>
          </div>
        </div>
      </div>`;
    }).join("");
  }

  let inited = false;
  _perfInit = function() {
    if (!inited) {
      buildCharts();
      document.querySelectorAll("#perfPresets button").forEach(b => {
        b.addEventListener("click", () => applyPerfPreset(b.dataset.d));
      });
      document.getElementById("perfFrom").addEventListener("change", e => {
        pstate.from = e.target.value;
        document.querySelectorAll("#perfPresets button").forEach(b => b.classList.remove("active"));
        rerender();
      });
      document.getElementById("perfTo").addEventListener("change", e => {
        pstate.to = e.target.value;
        document.querySelectorAll("#perfPresets button").forEach(b => b.classList.remove("active"));
        rerender();
      });
      const fromEl = document.getElementById("perfFrom"), toEl = document.getElementById("perfTo");
      if (PERF_MIN) { fromEl.min = PERF_MIN; toEl.min = PERF_MIN; }
      if (PERF_MAX) { fromEl.max = PERF_MAX; toEl.max = PERF_MAX; }
      applyPerfPreset("all");
      inited = true;
    }
    // Resize charts now that container is visible
    Object.values(pcharts).forEach(ch => { try { ch.resize(); ch.update(); } catch(e) {} });
  };
})();
</script>
</body>
</html>"""


def main(argv: list[str]) -> int:
    out_dir = Path(argv[1]) if len(argv) > 1 else Path(__file__).resolve().parent
    out_dir.mkdir(parents=True, exist_ok=True)
    src = out_dir / "source.xlsx"
    pcs_src = out_dir / "pcs_source.xlsx"
    html = out_dir / "kevin_dashboard.html"

    if not (len(argv) > 2 and argv[2] == "--no-download"):
        download_sheet(src, SHEET_URL)
        # PCS workbook: best-effort; if it fails we still render the dashboard
        # without the PCS column rather than blocking the refresh.
        try:
            download_sheet(pcs_src, PCS_SHEET_URL)
        except Exception as e:
            print(f"[etl] warning: PCS workbook download failed: {e}", file=sys.stderr)

    build(src, html, pcs_src if pcs_src.exists() else None)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
